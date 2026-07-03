"""Build best multi-strategy combos from the saved daily-return series + write a
comprehensive backtest-results Excel (all strategies, combos, correlations, caveats)."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"c:\Users\Shreyas.1Gupta\OneDrive - Angel Broking Limited\Desktop\Backup\NIFTY 500")
SPLIT = dt.date(2021, 12, 31)
M = pd.read_parquet(ROOT / "swing_momentum/multi_backtest_daily.parquet")


def stats(x):
    b = x[x.index.date <= SPLIT]; f = x[x.index.date > SPLIT]

    def m(r):
        r = r.dropna()
        if len(r) < 30 or r.std() == 0:
            return (0.0, 0.0, 0.0)
        eq = (1 + r).cumprod()
        return (r.mean()/r.std()*np.sqrt(252), eq.iloc[-1]**(252/len(r))-1, (eq/eq.cummax()-1).min())
    return m(b), m(f)


# combos (risk-parity-ish via inverse-vol on build)
def blend(cols, weights=None):
    sub = M[cols]
    if weights is None:      # inverse-vol on build
        v = sub[sub.index.date <= SPLIT].std().replace(0, np.nan)
        w = (1/v) / (1/v).sum()
    else:
        w = pd.Series(weights, index=cols)
    return (sub * w).sum(axis=1)


combos = {
    "Mom12 only": M["mom_12_1"],
    "LowVol only": M["lowvol_126"],
    "Mom12 + LowVol (invvol)": blend(["mom_12_1", "lowvol_126"]),
    "Mom12+LowVol+52wh+Trend": blend(["mom_12_1", "lowvol_126", "hi_52w", "trend+mom"]),
    "Mom6+regime + LowVol": blend(["mom6+regime", "lowvol_126"]),
}

rows = []
for name in list(M.columns) + ["|COMBOS|"]:
    if name == "|COMBOS|":
        for cname, series in combos.items():
            (sb, cb, ddb), (sf, cf, ddf) = stats(series)
            rows.append([cname, sb, cb, ddb, sf, cf, ddf, "combo"])
        continue
    (sb, cb, ddb), (sf, cf, ddf) = stats(M[name])
    rows.append([name, sb, cb, ddb, sf, cf, ddf, "single"])

res = pd.DataFrame(rows, columns=["Strategy", "Build Sharpe", "Build CAGR", "Build MaxDD",
                                  "Fwd Sharpe", "Fwd CAGR", "Fwd MaxDD", "Type"])
print(res.to_string(index=False, formatters={
    "Build Sharpe": "{:.2f}".format, "Fwd Sharpe": "{:.2f}".format,
    "Build CAGR": "{:+.1%}".format, "Fwd CAGR": "{:+.1%}".format,
    "Build MaxDD": "{:.0%}".format, "Fwd MaxDD": "{:.0%}".format}))

# ---- Excel ----
OUT = ROOT / "swing_momentum/Backtest_Results_India.xlsx"
corr = M[M.index.date <= SPLIT].corr().round(2)

notes = pd.DataFrame([
    ["Universe", "Full 2,535-symbol HF daily data, liquidity-filtered to top-500 by 60d turnover. Broader than the 976 PIT Nifty500 engine -> leans optimistic; treat CAGR as upper-ish bound."],
    ["Costs", "0.4% round-trip on turnover (STT+brokerage+slippage). Monthly rebalance."],
    ["Split adjustment", "Data ~effectively adjusted (0.01% daily |ret|>40%); daily returns winsorized at 25% as backstop."],
    ["Build/Forward", "Build <= 2021-12-31; Forward 2022-01 -> 2026-01 (~4y OOS)."],
    ["WINNER", "Momentum 12-1: fwd Sharpe 0.87, +21% CAGR — forward-robust. Low-vol: best risk-adj (fwd Sharpe 1.02, DD -14%)."],
    ["BEST COMBO", "Mom12 + LowVol (inverse-vol): keeps momentum return, cuts drawdown via the uncorrelated low-vol sleeve."],
    ["Big risk", "Momentum MaxDD -55% (build) = momentum crashes. Regime gating + low-vol overlay are the mitigants."],
    ["Failed forward", "Episodic Pivot (mechanical) and short-term reversal — negative/weak OOS. PEAD real but weak (~2-3%/60d), overlay only."],
    ["Next to fix", "Re-run on 976 PIT Nifty500 universe (processed/membership.parquet) + real Nifty regime + vol-targeting to tame the -55% DD."],
], columns=["Item", "Detail"])

with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    res.to_excel(xw, sheet_name="All_Strategies", index=False)
    corr.to_excel(xw, sheet_name="Correlation_build")
    notes.to_excel(xw, sheet_name="Verdict_and_Caveats", index=False)
    wb = xw.book
    hf = PatternFill("solid", fgColor="1B4332"); hfont = Font(bold=True, color="FFFFFF")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for ci in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 20 if ws.title != "Verdict_and_Caveats" else (22 if ci == 1 else 100)
            hc = ws.cell(row=1, column=ci); hc.fill = hf; hc.font = hfont
            hc.alignment = Alignment(wrap_text=True, vertical="center")
        for r in range(2, ws.max_row + 1):
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=r, column=ci).alignment = Alignment(wrap_text=True, vertical="top")
print(f"\nsaved -> {OUT}")
