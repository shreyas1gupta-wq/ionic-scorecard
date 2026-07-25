# -*- coding: utf-8 -*-
"""
mf_capture_recomm.py
=====================
Reverse-engineered, rerunnable recompute of the "MF Dashboard.xlsx" capture-ratio
Buy/Sell/Hold engine, for all 6 category pairs found in the workbook
(large, largemid, mid, flexi, multi, small -> "<cat>2" sheets carry the recommendation).

REVERSE-ENGINEERED LOGIC (traced from <cat>2 sheet formulas, verified against the
current cached workbook output for SMALL and FLEXI on 2026-07-25):

Raw sheet "<cat>" (e.g. "small"): daily NAVs, one column per fund, col A = date.
Benchmark: raw!DU2 names the category benchmark (pulled from "Indices" sheet via
INDEX/MATCH, approx/backward match on date).

"<cat>2" sheet re-samples raw NAVs onto MONTH-END anchor dates (col A) via
INDEX/MATCH(...,1) [= last raw daily row <= anchor date]. On top of those monthly
anchors it computes, per fund, per anchor row r (row offsets below are <cat>2's own
row spacing, 1 row = 1 month):

  FN  "Downside Capture (6M)"      = PRODUCT(1+fund daily ret, r-6..r, bench-down days only)-1
                                      / PRODUCT(1+bench daily ret, r-6..r, bench-down days only)-1
  HC  "Upside/Downside Capture(6M)"= [upside-capture(6M)] / [FN(6M)]   (upside computed the same
                                      way as FN but masked on bench-UP days)
        -> this HC ratio *is* the sheet's "Total Capture Ratio".
  IR  "Fund Ranking (with cutoff)" = IF(FN > LO1_threshold, exclude,
                                        RANK(HC, ALL funds' HC in the row, descending, ties=min))
        *** IMPORTANT: the RANK() comparison range is ALL funds, not just the funds
        that passed the FN cutoff. A fund can pass the cutoff yet still rank >3
        overall (see inconsistency #1 below). ***
  PK  "Quartile" (mislabeled: really a 4-way quadrant, not a statistical quartile)
                                   = 1 if HC>1 AND FN<1        (condition 2's branch is
                                                                 dead code: identical to
                                                                 condition 1, unreachable)
                                     3 if HC<1 AND FN>1
                                     4 otherwise (catch-all/default bucket)
  CJ  "12M Forward excess return"  = FundNAV(r+12)/FundNAV(r) - BenchNAV(r+12)/BenchNAV(r)
                                      (uses month-end NAV ratio directly, NOT daily-masked)
  MG/NV "1Y Downside/Total Capture" = same formula shape as FN/HC but windowed r-11..r
        *** mislabeled: this is an 11-MONTH window, not 12 (inconsistency #2). Not used
        by the recommendation itself (QZ references only FN/HC/PK/CJ, all 6M or 12M-true). ***
  QZ  "Recommendation"            = IF(IR<4, "BUY",
                                        IF(AND(CJ[r-12]<0, PK=4), "SELL", "HOLD"))
        where CJ[r-12] is the ALREADY-REALIZED trailing 12-month excess return ending at r
        (no lookahead: it's the CJ value computed 12 rows earlier, whose 12M forward window
        ends exactly at r).

Thresholds actually found in the workbook (cat2!LO1), differ from what was verbally
described by the Principal in one place -- see inconsistency #3 below:
  large2      LO1 = 0.9   benchmark = NIFTY 100
  largemid2   LO1 = 1.0   benchmark = NIFTY 250
  mid2        LO1 = 0.8   benchmark = NIFTY MIDCAP 150
  flexi2      LO1 = 1.0   benchmark = NIFTY 500
  multi2      LO1 = 0.9   benchmark = NIFTY MULTICAP 50:25:25   <-- Principal said 1.0
  small2      LO1 = 1.0   benchmark = NIFTY SMALLCAP 250

INCONSISTENCIES (flagged, not "corrected" -- this script reproduces the formulas AS
BUILT, bugs included, so it matches the live sheet):
  1. Buy-rank scope: IR's RANK() denominator is ALL funds' HC, not just cutoff
     survivors. A fund that passes the FN<=threshold filter is not guaranteed a
     top-3 rank even if it is the best among survivors, if non-surviving funds have
     higher HC. This is why, empirically, SMALL got only 2 BUYs and FLEXI only 1 BUY
     on 2025-01-31 instead of the "top 3" the methodology describes.
  2. PK ("Quartile") conditions 1 and 2 are IDENTICAL (AND(HC>1,FN<1) twice) -- branch
     2 is dead code. It is a quadrant classifier, not a statistical quartile.
  3. multi2's actual downside-capture cutoff is 0.9, not 1.0 as verbally described.
  4. MG/NV ("1Y" capture) windows are 11 calendar months (r-11..r), not 12 -- mislabeled.
     (Not load-bearing for the recommendation itself, which only uses FN/HC/PK/CJ.)
  5. KH1 ("Cutoff Rank-->" = 3) is a display label only -- IR/QZ hardcode the "<4"
     (top-3) cutoff directly and never reference KH1. Changing KH1 would silently do
     nothing.

Usage:
    python mf_capture_recomm.py                  # all 6 categories, latest anchor date
    python mf_capture_recomm.py --cats small flexi
    python mf_capture_recomm.py --verify          # also diff vs cached <cat>2!QZ output

Requires: openpyxl, pandas, numpy
"""
import argparse
import os
import shutil
import sys
import tempfile

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

ORIG_PATH = (
    r"C:\Users\Shreyas.1Gupta\OneDrive - Angel Broking Limited\Desktop\Backup"
    r"\NIFTY 500\MF Dashboard.xlsx"
)

# raw sheet name -> cat2 sheet name -> friendly label
CATEGORIES = {
    "small":    ("small2",    "SMALLCAP"),
    "flexi":    ("flexi2",    "FLEXICAP"),
    "large":    ("large2",    "LARGECAP"),
    "largemid": ("largemid2", "LARGE&MIDCAP"),
    "mid":      ("mid2",      "MIDCAP"),
    "multi":    ("multi2",    "MULTICAP"),
}


def resolve_path(path):
    """OneDrive caveat: if direct open fails, copy to %TEMP% first and open the copy."""
    try:
        with open(path, "rb"):
            pass
        return path
    except OSError:
        tmp = os.path.join(tempfile.gettempdir(), os.path.basename(path))
        shutil.copyfile(path, tmp)
        return tmp


def load_raw_sheet(wb_v, sheet_name, n_fund_cols_guess=60):
    """Return (dates: np.datetime64[], fund_names: list[str], nav: np.ndarray[T,F])."""
    ws = wb_v[sheet_name]
    # find contiguous fund columns from B onward using row2 header
    fund_names = []
    c = 2
    while True:
        v = ws.cell(row=2, column=c).value
        if v is None or v == "":
            break
        fund_names.append(v)
        c += 1
    n_funds = len(fund_names)
    last_col = 1 + n_funds  # col A + n_funds

    dates = []
    nav_rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=last_col, values_only=True):
        d = row[0]
        if not isinstance(d, (pd.Timestamp,)) and not hasattr(d, "year"):
            continue  # skips footer rows like "Inception Date"
        dates.append(pd.Timestamp(d))
        nav_rows.append([np.nan if v is None else float(v) for v in row[1:]])
    dates = pd.DatetimeIndex(dates)
    nav = np.array(nav_rows, dtype=float)
    return dates, fund_names, nav


def load_indices(wb_v):
    ws = wb_v["Indices"]
    names = []
    c = 2
    while True:
        v = ws.cell(row=2, column=c).value
        if v is None or v == "":
            break
        names.append(v)
        c += 1
    dates = []
    vals = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=1 + len(names), values_only=True):
        d = row[0]
        if d is None or not hasattr(d, "year"):
            continue
        dates.append(pd.Timestamp(d))
        vals.append([np.nan if v is None else float(v) for v in row[1:]])
    return pd.DatetimeIndex(dates), names, np.array(vals, dtype=float)


def benchmark_series_for(raw_dates, idx_dates, idx_vals_col):
    """Replicate INDEX(Indices!col,MATCH(date,Indices!A:A,1)) for every raw date:
    last Indices date <= raw date (approximate/backward match)."""
    pos = idx_dates.searchsorted(raw_dates, side="right") - 1
    pos = np.clip(pos, 0, len(idx_dates) - 1)
    return idx_vals_col[pos]


def month_end_anchors(dates):
    start = dates.min()
    end = dates.max()
    anchors = pd.date_range(start=start, end=end, freq="ME")
    return anchors


def anchor_positions(dates, anchors):
    """Position (0-based row index into `dates`) of the last date <= each anchor,
    replicating MATCH(anchor, raw!A:A, 1)."""
    pos = dates.searchsorted(anchors, side="right") - 1
    return pos


def excel_rank_desc_min(values):
    """Excel RANK(x, range) with default (descending) order, ties -> same (best) rank,
    next distinct value skips accordingly (== pandas rank(method='min', ascending=False))."""
    s = pd.Series(values)
    return s.rank(method="min", ascending=False)


def compute_category(wb_v, raw_name, cat2_name, label, target_anchor_idx=-1, verbose=True):
    ws2 = wb_v[cat2_name]
    lo1 = ws2.cell(row=1, column=column_index_from_string("LO")).value  # downside-capture cutoff
    kh1 = ws2.cell(row=1, column=column_index_from_string("KH")).value  # display-only, unused by formula

    dates, fund_names, nav = load_raw_sheet(wb_v, raw_name)
    bench_name = wb_v[raw_name].cell(row=2, column=column_index_from_string("DU")).value

    idx_dates, idx_names, idx_vals = load_indices(wb_v)
    bcol = idx_names.index(bench_name)
    bench_nav_daily = benchmark_series_for(dates, idx_dates, idx_vals[:, bcol])

    # daily returns (position-aligned with `dates`, row0 = NaN)
    fund_ret = np.vstack([np.nan * np.ones(nav.shape[1]), np.diff(nav, axis=0) / nav[:-1, :]])
    bench_ret = np.concatenate([[np.nan], np.diff(bench_nav_daily) / bench_nav_daily[:-1]])
    bench_up = (bench_ret > 0).astype(float)  # DY: 1 if bench return >0 else 0 (flat counts as "down")
    bench_up[np.isnan(bench_ret)] = np.nan

    anchors = month_end_anchors(dates)
    apos = anchor_positions(dates, anchors)

    r = len(anchors) + target_anchor_idx if target_anchor_idx < 0 else target_anchor_idx
    if r < 12:
        raise ValueError(f"{label}: not enough monthly history for a 12M-anchored calc")

    def window_prod(start_pos, end_pos, mask):
        """PRODUCT over daily rows [start_pos,end_pos] inclusive where mask==1;
        returns array over funds (nan-safe: skip both nan return and nan mask)."""
        seg_ret = fund_ret[start_pos:end_pos + 1, :]      # (D, F)
        seg_bmask = mask[start_pos:end_pos + 1]           # (D,)
        keep = (seg_bmask == 1)
        factor = np.where(keep[:, None] & ~np.isnan(seg_ret), 1.0 + seg_ret, 1.0)
        factor[np.isnan(seg_ret)] = 1.0  # blanks skipped, same as Excel PRODUCT ignoring text/blank
        # but cells outside the mask must NOT be multiplied in at all (equiv factor=1) - matches
        return factor.prod(axis=0)

    def bench_window_prod(start_pos, end_pos, mask):
        seg_ret = bench_ret[start_pos:end_pos + 1]
        seg_bmask = mask[start_pos:end_pos + 1]
        keep = (seg_bmask == 1) & ~np.isnan(seg_ret)
        factor = np.where(keep, 1.0 + seg_ret, 1.0)
        return factor.prod()

    s6, e6 = apos[r - 6], apos[r]
    down_mask = 1 - bench_up
    up_mask = bench_up

    fund_down_prod = window_prod(s6, e6, down_mask)
    bench_down_prod = bench_window_prod(s6, e6, down_mask)
    FN = (fund_down_prod - 1) / (bench_down_prod - 1)  # Downside Capture (6M)

    fund_up_prod = window_prod(s6, e6, up_mask)
    bench_up_prod = bench_window_prod(s6, e6, up_mask)
    upside = (fund_up_prod - 1) / (bench_up_prod - 1)
    HC = upside / FN  # "Total Capture Ratio" = Upside Capture / Downside Capture (6M)

    # PK: quadrant (mislabeled "Quartile")
    PK = np.where((HC > 1) & (FN < 1), 1, np.where((HC < 1) & (FN > 1), 3, 4)).astype(float)
    PK[np.isnan(HC) | np.isnan(FN)] = np.nan

    # IR: rank by HC among ALL funds (not just survivors), then blank out non-survivors
    ranks = excel_rank_desc_min(HC).to_numpy()
    IR = np.where(FN > lo1, np.nan, ranks)

    # CJ at r-12: trailing 12M excess return ending at r (month-end NAV ratio, no daily masking)
    pos_r = apos[r]
    pos_rm12 = apos[r - 12]
    fund_nav_r = nav[pos_r, :]
    fund_nav_rm12 = nav[pos_rm12, :]
    bench_nav_r = bench_nav_daily[pos_r]
    bench_nav_rm12 = bench_nav_daily[pos_rm12]
    CJ_rm12 = (fund_nav_r / fund_nav_rm12 - 1) - (bench_nav_r / bench_nav_rm12 - 1)

    # QZ recommendation
    reco = []
    for i in range(len(fund_names)):
        if np.isnan(nav[pos_r, i]):
            reco.append("")
            continue
        if not np.isnan(IR[i]) and IR[i] < 4:
            reco.append("BUY")
        elif (not np.isnan(CJ_rm12[i]) and CJ_rm12[i] < 0) and (not np.isnan(PK[i]) and PK[i] == 4):
            reco.append("SELL")
        else:
            reco.append("HOLD")

    result = pd.DataFrame({
        "fund": fund_names,
        "FN_downside_cap_6M": FN,
        "HC_total_cap_6M": HC,
        "IR_rank": IR,
        "PK_quadrant": PK,
        "CJ_trailing12M_excess": CJ_rm12,
        "recommendation": reco,
    })

    if verbose:
        print(f"\n=== {label}  (raw={raw_name}, cat2={cat2_name}) ===")
        print(f"benchmark={bench_name}  6M-downside-cutoff(LO1)={lo1}  "
              f"display-only rank-cutoff(KH1)={kh1}  as-of anchor={anchors[r].date()}")
        buys = result[result.recommendation == "BUY"].sort_values("IR_rank")
        sells = result[result.recommendation == "SELL"]
        print(f"BUY ({len(buys)}):")
        for _, row in buys.iterrows():
            print(f"   rank {int(row.IR_rank)}: {row.fund}  (FN={row.FN_downside_cap_6M:.3f}, "
                  f"HC={row.HC_total_cap_6M:.3f})")
        print(f"SELL ({len(sells)}):")
        for _, row in sells.iterrows():
            print(f"   {row.fund}  (PK=4, CJ_trail12M_excess={row.CJ_trailing12M_excess:+.3%})")
        n_hold = (result.recommendation == "HOLD").sum()
        print(f"HOLD: {n_hold} funds")

    return result, anchors[r]


def verify_against_sheet(wb_v, cat2_name, anchor_date, computed_df):
    ws2 = wb_v[cat2_name]
    # find the row in cat2 whose col-A date == anchor_date
    row_num = None
    for r in range(7, ws2.max_row + 1):
        v = ws2.cell(row=r, column=1).value
        if v is not None and pd.Timestamp(v) == pd.Timestamp(anchor_date):
            row_num = r
            break
    if row_num is None:
        print(f"  [verify] could not locate anchor {anchor_date.date()} in {cat2_name}!A -- skipped")
        return
    qz_start = column_index_from_string("QZ")
    n = len(computed_df)
    cached = []
    for i in range(n):
        c = qz_start + i
        fname = ws2.cell(row=5, column=c).value
        rec = ws2.cell(row=row_num, column=c).value
        cached.append((fname, rec))
    cached_df = pd.DataFrame(cached, columns=["fund", "cached_recommendation"])
    merged = computed_df.merge(cached_df, on="fund", how="left")
    merged["match"] = merged["recommendation"].fillna("") == merged["cached_recommendation"].fillna("")
    mism = merged[~merged["match"]]
    print(f"  [verify] row {row_num} ({anchor_date.date()}): "
          f"{merged['match'].sum()}/{len(merged)} funds match cached QZ output")
    if len(mism):
        print("  MISMATCHES:")
        for _, row in mism.iterrows():
            print(f"    {row.fund}: computed={row.recommendation!r} vs cached={row.cached_recommendation!r}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", default=ORIG_PATH)
    ap.add_argument("--cats", nargs="*", default=list(CATEGORIES.keys()),
                     help="raw sheet names to run, e.g. small flexi")
    ap.add_argument("--verify", action="store_true",
                     help="also diff computed recommendations vs the sheet's cached QZ output")
    args = ap.parse_args()

    path = resolve_path(args.path)
    wb_v = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for raw_name in args.cats:
        if raw_name not in CATEGORIES:
            print(f"unknown category '{raw_name}', skipping", file=sys.stderr)
            continue
        cat2_name, label = CATEGORIES[raw_name]
        df, anchor_date = compute_category(wb_v, raw_name, cat2_name, label)
        if args.verify:
            verify_against_sheet(wb_v, cat2_name, anchor_date, df)


if __name__ == "__main__":
    main()
