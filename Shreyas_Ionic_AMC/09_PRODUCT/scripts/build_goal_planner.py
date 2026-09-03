# -*- coding: utf-8 -*-
"""Builds the Ionic goal planning workbook.

Same six sheets as the version it replaces, same five worked example goals, with the
return ladder, the priority mapping and the erosion column corrected. Every bucket
return is a live Excel formula on two named planning inputs rather than a typed number.

Run with no arguments. Output goes to 09_PRODUCT/reports.
"""
import math
import os
import re
import zipfile
from xml.sax.saxutils import escape

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.abspath(os.path.join(HERE, "..", "reports", "Ionic_Goal_Planner.xlsx"))

# ---------------------------------------------------------------------------
# The two planning inputs. Everything else is derived.
# ---------------------------------------------------------------------------
EQUITY = 0.110
DEBT = 0.060

MONEY = "#,##0"
BIGMONEY = '"Rs "#,##0'
PCT1 = "0.0%"
PCT2 = "0.00%"
YEARS = "0"

INK = "1F3864"
BAND = "2E5496"
SOFT = "D9E2F3"
PALE = "F2F5FB"
LINE = "B4C6E7"

TITLEFONT = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
HEADFONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=10, bold=True)
BODY = Font(name="Calibri", size=10)
SMALL = Font(name="Calibri", size=9, color="404040")
SMALLITAL = Font(name="Calibri", size=9, italic=True, color="404040")

TITLEFILL = PatternFill("solid", fgColor=INK)
HEADFILL = PatternFill("solid", fgColor=BAND)
INFILL = PatternFill("solid", fgColor=SOFT)
PALEFILL = PatternFill("solid", fgColor=PALE)

THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAPTOP = Alignment(wrap_text=True, vertical="top")
WRAPMID = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ---------------------------------------------------------------------------
# Buckets, ordered from most equity to least. Priority now sets risk downward:
# a high priority goal is one that cannot be missed, so it takes the least equity.
# No bucket inside two years holds any equity at all.
# ---------------------------------------------------------------------------
BUCKETS = [
    ("L-L", "Long-term Aggressive", "Long (5+ yrs)", "Low", 1.00,
     "Diversified equity funds and index funds"),
    ("L-M", "Long-term Growth", "Long (5+ yrs)", "Medium", 0.75,
     "Equity funds with a short duration debt holding"),
    ("L-H", "Long-term Conservative", "Long (5+ yrs)", "High", 0.50,
     "Multi asset funds and corporate bond funds"),
    ("M-L", "Medium-term Growth", "Medium (2-5 yrs)", "Low", 0.60,
     "Aggressive hybrid funds and short duration debt"),
    ("M-M", "Medium-term Moderate", "Medium (2-5 yrs)", "Medium", 0.40,
     "Balanced advantage funds and corporate bond funds"),
    ("M-H", "Medium-term Conservative", "Medium (2-5 yrs)", "High", 0.20,
     "Conservative hybrid funds and short duration debt"),
    ("S-L", "Short-term Accrual", "Short (0-2 yrs)", "Low", 0.00,
     "Short duration debt funds"),
    ("S-M", "Short-term Moderate", "Short (0-2 yrs)", "Medium", 0.00,
     "Ultra short duration and money market funds"),
    ("S-H", "Short-term Capital Protection", "Short (0-2 yrs)", "High", 0.00,
     "Liquid funds and arbitrage funds"),
]
BUCKETROW0 = 6
EQUITYWEIGHT = {b[1]: b[4] for b in BUCKETS}

INFLATION = [
    ("Child's Education (Domestic)", 0.10, 0.12, 0.11,
     "Private institution fees and campus costs rise at close to double the rate of "
     "general consumer inflation."),
    ("Child's Education (Overseas)", 0.12, 0.14, 0.13,
     "As above, with currency depreciation added on tuition and living costs abroad."),
    ("Emergency Fund", 0.06, 0.06, 0.06,
     "Covers monthly household spending at consumer price inflation, so the fund keeps "
     "its purchasing power."),
    ("Vacation Planning", 0.08, 0.10, 0.09,
     "Air fares and hotel tariffs run ahead of general retail price increases."),
    ("Entrepreneurial Venture", 0.08, 0.10, 0.09,
     "Commercial rent and skilled salary costs climb faster than general consumer goods."),
    ("Retirement Corpus", 0.06, 0.07, 0.065,
     "Post retirement spending mixes household costs with medical costs, which run "
     "hotter. A blended rate is used."),
    ("Other / Custom", 0.07, 0.07, 0.07,
     "General consumer inflation placeholder. Override the rate on the goal row when "
     "you have something better for that goal."),
]
INFLROW0 = 6
MIDRATE = {r[0]: r[3] for r in INFLATION}

GOALS = [
    ("Child's Education", "Child's Education (Domestic)", 3000000, 4, "High"),
    ("Emergency Fund", "Emergency Fund", 1000000, 1, "Medium"),
    ("Vacation Planning", "Vacation Planning", 400000, 2, "Low"),
    ("Entrepreneurial Venture", "Entrepreneurial Venture", 25000000, 6, "High"),
    ("Retirement Corpus", "Retirement Corpus", 40000000, 25, "High"),
]
GOALROW0 = 5
GOALROWS = 10

# What the version being replaced used, kept only so the change can be priced.
PRIORRATE = [0.15, 0.08, 0.06, 0.18, 0.18]
PRIORWEIGHT = [0.70, 0.10, 0.00, 1.00, 1.00]

NOTPROJECTION = ("These are planning assumptions. They are not a projection, not a forecast, "
                 "not a guarantee and not a promise of return. Actual returns will differ from "
                 "them, and over any horizon shown here they can be negative.")
SERIESLIMIT = ("The equity series behind the equity input is 20.9 years long. The retirement "
               "goal runs 25 years, so no 25 year holding period can be measured from it at "
               "all. Any 25 year figure in this workbook is an extrapolation beyond the "
               "evidence, and it is shown because the goal exists, not because the number has "
               "been tested.")
PRIORITYRULE = ("Priority sets risk downward, not upward. A high priority goal is the one you "
                "cannot afford to miss, so it takes the least equity. A low priority goal can "
                "slip a year or be spent smaller, so it can take the most.")


def bucketfor(years, priority):
    if years <= 2:
        band = "Short"
    elif years <= 5:
        band = "Medium"
    else:
        band = "Long"
    table = {
        ("Short", "High"): "Short-term Capital Protection",
        ("Short", "Medium"): "Short-term Moderate",
        ("Short", "Low"): "Short-term Accrual",
        ("Medium", "High"): "Medium-term Conservative",
        ("Medium", "Medium"): "Medium-term Moderate",
        ("Medium", "Low"): "Medium-term Growth",
        ("Long", "High"): "Long-term Conservative",
        ("Long", "Medium"): "Long-term Growth",
        ("Long", "Low"): "Long-term Aggressive",
    }
    return table[(band, priority)]


def bucketreturn(name):
    w = EQUITYWEIGHT[name]
    return w * EQUITY + (1.0 - w) * DEBT


class Book(object):
    """Wraps openpyxl so that every formula carries the value it should evaluate to.

    openpyxl cannot calculate, so the values are injected into the saved file as the
    cached results Excel stores next to a formula. Excel recalculates on open, which
    means a cached value can never quietly outlive the formula that produced it.
    """

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.cache = {}

    def sheet(self, name):
        ws = self.wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        self.cache[name] = {}
        return ws

    def val(self, ws, coord, value, fmt=None, font=BODY, fill=None, align=None, box=False):
        c = ws[coord]
        c.value = value
        c.font = font
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if box:
            c.border = BOX
        return c

    def formula(self, ws, coord, text, cached, fmt=None, font=BODY, fill=None,
                align=None, box=False):
        c = self.val(ws, coord, text, fmt=fmt, font=font, fill=fill, align=align, box=box)
        self.cache[ws.title][coord] = cached
        return c

    def title(self, ws, row, lastcol, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=lastcol)
        self.val(ws, "A%d" % row, text, font=TITLEFONT, fill=TITLEFILL,
                 align=Alignment(vertical="center"))
        ws.row_dimensions[row].height = 24

    def note(self, ws, row, lastcol, text, height=15, font=SMALLITAL):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=lastcol)
        self.val(ws, "A%d" % row, text, font=font, align=WRAPTOP)
        ws.row_dimensions[row].height = height

    def header(self, ws, row, labels, height=32):
        for i, label in enumerate(labels, start=1):
            self.val(ws, "%s%d" % (get_column_letter(i), row), label,
                     font=HEADFONT, fill=HEADFILL, align=WRAPMID, box=True)
        ws.row_dimensions[row].height = height

    def widths(self, ws, spec):
        for col, w in spec.items():
            ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Cached value injection
# ---------------------------------------------------------------------------
CELLRE = re.compile(r'<c r="([A-Z]+[0-9]+)"([^>]*)><f([^>]*)>(.*?)</f><v></v></c>', re.DOTALL)


def injectcached(path, cache):
    zin = zipfile.ZipFile(path)
    infos = zin.infolist()
    blobs = {i.filename: zin.read(i.filename) for i in infos}
    zin.close()

    rels = blobs["xl/_rels/workbook.xml.rels"].decode("utf-8")
    idmap = {}
    for chunk in re.findall(r"<Relationship\b[^>]*/>", rels):
        rid = re.search(r'Id="([^"]+)"', chunk)
        tgt = re.search(r'Target="([^"]+)"', chunk)
        if rid and tgt:
            idmap[rid.group(1)] = tgt.group(1)

    wbxml = blobs["xl/workbook.xml"].decode("utf-8")
    targets = {}
    for chunk in re.findall(r"<sheet\b[^>]*/>", wbxml):
        nm = re.search(r'name="([^"]+)"', chunk)
        rid = re.search(r'r:id="([^"]+)"', chunk)
        if not nm or not rid:
            continue
        tgt = idmap[rid.group(1)].replace("\\", "/").lstrip("/")
        if not tgt.startswith("xl/"):
            tgt = "xl/" + tgt
        targets[nm.group(1).replace("&amp;", "&")] = tgt

    written = 0
    for name, values in cache.items():
        target = targets[name]
        xml = blobs[target].decode("utf-8")

        def repl(m):
            ref, attrs, fattrs, body = m.group(1), m.group(2), m.group(3), m.group(4)
            if ref not in values:
                return m.group(0)
            v = values[ref]
            if v is None or v == "":
                return m.group(0)
            attrs = re.sub(r'\s+t="[^"]*"', "", attrs)
            if isinstance(v, bool):
                return '<c r="%s"%s t="b"><f%s>%s</f><v>%d</v></c>' % (
                    ref, attrs, fattrs, body, 1 if v else 0)
            if isinstance(v, (int, float)):
                return '<c r="%s"%s><f%s>%s</f><v>%s</v></c>' % (
                    ref, attrs, fattrs, body, repr(float(v)))
            return '<c r="%s"%s t="str"><f%s>%s</f><v>%s</v></c>' % (
                ref, attrs, fattrs, body, escape(str(v)))

        xml, n = CELLRE.subn(repl, xml)
        written += n
        blobs[target] = xml.encode("utf-8")

    zout = zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED)
    for i in infos:
        zout.writestr(i, blobs[i.filename])
    zout.close()
    return written


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------
def goalsinput(bk, model):
    ws = bk.sheet("Goals Input")
    bk.widths(ws, {"A": 22, "B": 27, "C": 16, "D": 13, "E": 11, "F": 12, "G": 19,
                   "H": 17, "I": 27, "J": 13, "K": 19, "L": 14, "M": 16})
    bk.title(ws, 1, 13, "IONIC WEALTH GOAL PLANNING TOOL")
    bk.note(ws, 2, 13,
            "Fill the shaded cells: goal name, goal category, cost in today's money, "
            "horizon in years, priority. The rest of the row calculates itself, including "
            "the inflation adjustment, the bucket and the amount the goal needs today.",
            height=28)
    bk.note(ws, 3, 13, PRIORITYRULE, height=28, font=Font(name="Calibri", size=9, bold=True,
                                                          color="1F3864"))
    bk.header(ws, 4, [
        "Goal Name", "Goal Category", "Today's Cost (Rs)", "Time Horizon (yrs)", "Priority",
        "Inflation Rate Applied", "Inflation-Adjusted Future Value (Rs)",
        "Cost Increase (Rs)", "Bucket Mapped", "Expected Return (Bucket)",
        "PV Required Today (Rs)", "Cost Increase (%)", "Purchasing Power Lost (%)"])

    ilo = INFLROW0
    ihi = INFLROW0 + len(INFLATION) - 1
    blo = BUCKETROW0
    bhi = BUCKETROW0 + len(BUCKETS) - 1

    for i in range(GOALROWS):
        r = GOALROW0 + i
        filled = i < len(GOALS)
        g = model["goals"][i] if filled else None

        for col in ("A", "B", "C", "D", "E"):
            bk.val(ws, "%s%d" % (col, r), None, fill=INFILL, box=True)
        if filled:
            bk.val(ws, "A%d" % r, g["name"], font=BOLD, fill=INFILL, box=True)
            bk.val(ws, "B%d" % r, g["category"], fill=INFILL, box=True)
            bk.val(ws, "C%d" % r, g["cost"], fmt=MONEY, fill=INFILL, box=True)
            bk.val(ws, "D%d" % r, g["years"], fmt=YEARS, fill=INFILL, box=True,
                   align=Alignment(horizontal="center"))
            bk.val(ws, "E%d" % r, g["priority"], fill=INFILL, box=True,
                   align=Alignment(horizontal="center"))

        bk.formula(ws, "F%d" % r,
                   "=IFERROR(INDEX('Inflation Assumptions'!$D$%d:$D$%d, "
                   "MATCH(B%d, 'Inflation Assumptions'!$A$%d:$A$%d, 0)), \"\")"
                   % (ilo, ihi, r, ilo, ihi),
                   g["infl"] if filled else "", fmt=PCT1, box=True)
        bk.formula(ws, "G%d" % r,
                   "=IF(AND(C%d<>\"\", D%d<>\"\", F%d<>\"\"), C%d*POWER(1+F%d, D%d), \"\")"
                   % (r, r, r, r, r, r),
                   g["fv"] if filled else "", fmt=MONEY, box=True)
        bk.formula(ws, "H%d" % r, "=IF(G%d<>\"\", G%d-C%d, \"\")" % (r, r, r),
                   g["fv"] - g["cost"] if filled else "", fmt=MONEY, box=True)
        bk.formula(ws, "I%d" % r,
                   "=IF(AND(D{r}<>\"\",E{r}<>\"\"),"
                   "IF(D{r}<=2,IF(E{r}=\"High\",\"Short-term Capital Protection\","
                   "IF(E{r}=\"Medium\",\"Short-term Moderate\",\"Short-term Accrual\")),"
                   "IF(D{r}<=5,IF(E{r}=\"High\",\"Medium-term Conservative\","
                   "IF(E{r}=\"Medium\",\"Medium-term Moderate\",\"Medium-term Growth\")),"
                   "IF(E{r}=\"High\",\"Long-term Conservative\","
                   "IF(E{r}=\"Medium\",\"Long-term Growth\",\"Long-term Aggressive\")))),"
                   "\"\")".format(r=r),
                   g["bucket"] if filled else "", box=True)
        bk.formula(ws, "J%d" % r,
                   "=IFERROR(INDEX(Assumptions!$G$%d:$G$%d, "
                   "MATCH(I%d, Assumptions!$B$%d:$B$%d, 0)), \"\")" % (blo, bhi, r, blo, bhi),
                   g["ret"] if filled else "", fmt=PCT2, box=True)
        bk.formula(ws, "K%d" % r,
                   "=IF(AND(G%d<>\"\", J%d<>\"\"), G%d/POWER(1+J%d, D%d), \"\")"
                   % (r, r, r, r, r),
                   g["pv"] if filled else "", fmt=MONEY, font=BOLD, box=True)
        bk.formula(ws, "L%d" % r,
                   "=IF(AND(H%d<>\"\", C%d<>0, C%d<>\"\"), H%d/C%d, \"\")" % (r, r, r, r, r),
                   g["costpct"] if filled else "", fmt=PCT2, box=True)
        bk.formula(ws, "M%d" % r,
                   "=IF(AND(F%d<>\"\", D%d<>\"\"), 1-1/POWER(1+F%d, D%d), \"\")"
                   % (r, r, r, r),
                   g["lost"] if filled else "", fmt=PCT2, box=True)

    last = GOALROW0 + GOALROWS - 1
    bk.val(ws, "A%d" % (last + 1), "Plan total", font=BOLD)
    bk.formula(ws, "G%d" % (last + 1), "=SUM(G%d:G%d)" % (GOALROW0, last),
               model["totalfv"], fmt=MONEY, font=BOLD)
    bk.formula(ws, "K%d" % (last + 1), "=SUM(K%d:K%d)" % (GOALROW0, last),
               model["totalpv"], fmt=MONEY, font=BOLD)

    bk.note(ws, last + 3, 13,
            "Cost Increase (%) is the future cost divided by today's cost, less one. It is a "
            "gross up ratio and it has no ceiling. Purchasing Power Lost (%) is one minus one "
            "divided by (1 plus inflation) raised to the horizon. It is bounded at 100% by "
            "construction, so any figure above 100% in that column would be arithmetically "
            "impossible.", height=42)
    bk.note(ws, last + 4, 13, SERIESLIMIT, height=42)
    bk.note(ws, last + 5, 13, NOTPROJECTION, height=30,
            font=Font(name="Calibri", size=9, bold=True, color="C00000"))

    dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("E%d:E%d" % (GOALROW0, last))
    dvc = DataValidation(
        type="list",
        formula1="'Inflation Assumptions'!$A$%d:$A$%d" % (ilo, ihi), allow_blank=True)
    ws.add_data_validation(dvc)
    dvc.add("B%d:B%d" % (GOALROW0, last))

    ws.freeze_panes = "A%d" % GOALROW0
    return ws


def inflationsheet(bk):
    ws = bk.sheet("Inflation Assumptions")
    bk.widths(ws, {"A": 28, "B": 10, "C": 10, "D": 20, "E": 78})
    bk.title(ws, 1, 5, "GOAL CATEGORY INFLATION ASSUMPTIONS")
    bk.note(ws, 3, 5,
            "These rates carry a goal's cost forward from today's prices. The recommended mid "
            "rate is the one the goal rows look up. Override the rate on a goal row when a "
            "goal has its own known cost path.", height=28)
    bk.header(ws, 5, ["Goal Category", "Low %", "High %", "Recommended (Mid) %",
                      "Why this rate"])
    for i, (cat, lo, hi, mid, why) in enumerate(INFLATION):
        r = INFLROW0 + i
        bk.val(ws, "A%d" % r, cat, font=BOLD, box=True)
        bk.val(ws, "B%d" % r, lo, fmt=PCT1, box=True)
        bk.val(ws, "C%d" % r, hi, fmt=PCT1, box=True)
        bk.val(ws, "D%d" % r, mid, fmt=PCT1, box=True, fill=PALEFILL, font=BOLD)
        bk.val(ws, "E%d" % r, why, box=True, align=WRAPTOP)
        ws.row_dimensions[r].height = 30
    bk.note(ws, INFLROW0 + len(INFLATION) + 1, 5, NOTPROJECTION, height=28)
    ws.freeze_panes = "A6"
    return ws


def assumptions(bk, model):
    ws = bk.sheet("Assumptions")
    bk.widths(ws, {"A": 13, "B": 30, "C": 18, "D": 11, "E": 13, "F": 12, "G": 15,
                   "H": 22, "I": 46})
    bk.title(ws, 1, 9, "BUCKET ASSUMPTIONS AND MAPPING")

    bk.val(ws, "A2", "Equity planning return", font=BOLD)
    bk.val(ws, "B2", EQUITY, fmt=PCT2, font=BOLD, fill=INFILL, box=True,
           align=Alignment(horizontal="center"))
    bk.val(ws, "C2", "Debt planning return", font=BOLD)
    bk.val(ws, "D2", DEBT, fmt=PCT2, font=BOLD, fill=INFILL, box=True,
           align=Alignment(horizontal="center"))
    ws.merge_cells("E2:I2")
    bk.val(ws, "E2",
           "Change either cell and every bucket return, every goal and every total in this "
           "workbook moves with it.", font=SMALLITAL, align=WRAPTOP)

    bk.note(ws, 3, 9,
            "Expected Return below is calculated from each bucket's own equity and debt "
            "weights and the two cells above. No return in this table is typed in by hand, "
            "which is what stopped the same mix being priced three different ways.", height=28)

    bk.header(ws, 5, ["Bucket Code", "Bucket Name", "Time Horizon", "Priority",
                      "Equity Weight", "Debt Weight", "Expected Return",
                      "Equity : Debt Mix", "Asset Classes"])
    for i, (code, name, band, priority, w, assets) in enumerate(BUCKETS):
        r = BUCKETROW0 + i
        bk.val(ws, "A%d" % r, code, box=True, align=Alignment(horizontal="center"))
        bk.val(ws, "B%d" % r, name, font=BOLD, box=True)
        bk.val(ws, "C%d" % r, band, box=True)
        bk.val(ws, "D%d" % r, priority, box=True, align=Alignment(horizontal="center"))
        bk.val(ws, "E%d" % r, w, fmt=PCT1, box=True, fill=INFILL)
        bk.formula(ws, "F%d" % r, "=1-E%d" % r, 1.0 - w, fmt=PCT1, box=True)
        bk.formula(ws, "G%d" % r, "=E%d*EquityReturn+F%d*DebtReturn" % (r, r),
                   w * EQUITY + (1.0 - w) * DEBT, fmt=PCT2, font=BOLD, box=True,
                   fill=PALEFILL)
        bk.formula(ws, "H%d" % r,
                   '=TEXT(E%d,"0%%")&" equity / "&TEXT(F%d,"0%%")&" debt"' % (r, r),
                   "%d%% equity / %d%% debt" % (round(w * 100), round((1 - w) * 100)),
                   box=True)
        bk.val(ws, "I%d" % r, assets, box=True, align=WRAPTOP)
        ws.row_dimensions[r].height = 26

    r = BUCKETROW0 + len(BUCKETS) + 1
    notes = [
        ("Where the equity input comes from",
         "The NIFTY 50 compounded 13.02% on price over 20.8 years, but that window opens in "
         "April 2005, close to the base of a bull run, and a start date chosen there flatters "
         "the answer. Measured from the January 2008 peak the same index compounded 8.27% on "
         "price over 18 years. The 15 year rolling median is about 10.8% on price, and a "
         "measured dividend yield of about 1.27% takes that median to about 12.1% total "
         "return. The planning input is set at 11.0%, below the median, because a plan built "
         "on the median misses half the time."),
        ("Where the debt input comes from",
         "6.0% is carried over unchanged from the earlier version, which used it for its "
         "all debt bucket. It is the one input in that version that did not need correcting."),
        ("How priority now works", PRIORITYRULE + " The earlier version did the reverse and "
         "put the most equity behind the goals with the least room to fail."),
        ("Why nothing inside two years holds equity",
         "Two years does not give an equity holding time to recover from a drawdown, and a "
         "goal that lands inside a drawdown is a goal that is not funded. All three short "
         "horizon buckets are debt only and all three earn the debt input. They differ by "
         "the duration of the debt they hold, not by how much risk they carry."),
        ("What this evidence cannot support", SERIESLIMIT),
        ("Status of these numbers", NOTPROJECTION),
    ]
    for label, text in notes:
        bk.val(ws, "A%d" % r, label, font=BOLD, align=WRAPTOP)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        bk.val(ws, "B%d" % r, text, font=SMALL, align=WRAPTOP)
        ws.row_dimensions[r].height = max(30, 13 * math.ceil(len(text) / 150.0) + 8)
        r += 1

    ws.freeze_panes = "A6"
    return ws


def calculations(bk, model):
    ws = bk.sheet("Calculations")
    bk.widths(ws, {"A": 42, "B": 18, "C": 22, "D": 16})
    bk.title(ws, 1, 4, "PLAN CALCULATIONS")
    lastgoal = GOALROW0 + GOALROWS - 1

    rows = [
        ("Total future value required (inflation adjusted):",
         "=SUM('Goals Input'!G%d:G%d)" % (GOALROW0, lastgoal), model["totalfv"], BIGMONEY),
        ("Total cost increase over the plan (Rs):",
         "=SUM('Goals Input'!H%d:H%d)" % (GOALROW0, lastgoal),
         model["totalfv"] - model["totalcost"], BIGMONEY),
        ("Total amount required today (Rs):",
         "=SUM('Goals Input'!K%d:K%d)" % (GOALROW0, lastgoal), model["totalpv"], BIGMONEY),
        ("Number of goals:",
         "=COUNTA('Goals Input'!A%d:A%d)" % (GOALROW0, lastgoal), len(GOALS), YEARS),
    ]
    for i, (label, fml, cached, fmt) in enumerate(rows):
        r = 3 + i
        bk.val(ws, "A%d" % r, label, font=BOLD)
        bk.formula(ws, "B%d" % r, fml, cached, fmt=fmt, font=BOLD, box=True, fill=PALEFILL)

    bk.val(ws, "A8", "Bucket distribution", font=BOLD)
    bk.header(ws, 9, ["Bucket", "Goals", "Amount required today (Rs)", "% of plan"])
    for i, b in enumerate(BUCKETS):
        r = 10 + i
        name = b[1]
        bk.val(ws, "A%d" % r, name, box=True)
        bk.formula(ws, "B%d" % r,
                   "=COUNTIF('Goals Input'!$I$%d:$I$%d, $A%d)" % (GOALROW0, lastgoal, r),
                   model["bucketcount"][name], fmt=YEARS, box=True,
                   align=Alignment(horizontal="center"))
        bk.formula(ws, "C%d" % r,
                   "=SUMIF('Goals Input'!$I$%d:$I$%d, $A%d, 'Goals Input'!$K$%d:$K$%d)"
                   % (GOALROW0, lastgoal, r, GOALROW0, lastgoal),
                   model["bucketpv"][name], fmt=MONEY, box=True)
        bk.formula(ws, "D%d" % r, "=IFERROR(C%d/$C$%d, 0)" % (r, 10 + len(BUCKETS)),
                   model["bucketpv"][name] / model["totalpv"], fmt=PCT1, box=True)
    tot = 10 + len(BUCKETS)
    bk.val(ws, "A%d" % tot, "Total", font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "B%d" % tot, "=SUM(B10:B%d)" % (tot - 1), len(GOALS), fmt=YEARS,
               font=BOLD, box=True, fill=PALEFILL, align=Alignment(horizontal="center"))
    bk.formula(ws, "C%d" % tot, "=SUM(C10:C%d)" % (tot - 1), model["totalpv"], fmt=MONEY,
               font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "D%d" % tot, "=SUM(D10:D%d)" % (tot - 1), 1.0, fmt=PCT1, font=BOLD,
               box=True, fill=PALEFILL)
    ws.freeze_panes = "A10"
    return ws


def allocation(bk, model):
    ws = bk.sheet("Portfolio Allocation")
    bk.widths(ws, {"A": 30, "B": 20, "C": 13, "D": 14, "E": 18, "F": 18, "G": 46})
    bk.title(ws, 1, 7, "RECOMMENDED PORTFOLIO STRUCTURE")
    bk.note(ws, 2, 7,
            "Amounts are what each bucket needs today. The equity and debt split is read "
            "from the bucket weights on the Assumptions sheet, so the plan's overall equity "
            "share is a result of the goals, not a number set separately.", height=28)
    bk.header(ws, 3, ["Bucket", "Amount today (Rs)", "% of plan", "Equity weight",
                      "Equity (Rs)", "Debt (Rs)", "Asset Classes"])
    lastgoal = GOALROW0 + GOALROWS - 1
    blo, bhi = BUCKETROW0, BUCKETROW0 + len(BUCKETS) - 1
    first, last = 4, 4 + len(BUCKETS) - 1
    for i, b in enumerate(BUCKETS):
        r = first + i
        name = b[1]
        amt = model["bucketpv"][name]
        w = b[4]
        bk.val(ws, "A%d" % r, name, font=BOLD, box=True)
        bk.formula(ws, "B%d" % r,
                   "=SUMIF('Goals Input'!$I$%d:$I$%d, $A%d, 'Goals Input'!$K$%d:$K$%d)"
                   % (GOALROW0, lastgoal, r, GOALROW0, lastgoal), amt, fmt=MONEY, box=True)
        bk.formula(ws, "C%d" % r, "=IFERROR(B%d/$B$%d, 0)" % (r, last + 1),
                   amt / model["totalpv"], fmt=PCT1, box=True)
        bk.formula(ws, "D%d" % r,
                   "=IFERROR(INDEX(Assumptions!$E$%d:$E$%d, MATCH($A%d, "
                   "Assumptions!$B$%d:$B$%d, 0)), 0)" % (blo, bhi, r, blo, bhi),
                   w, fmt=PCT1, box=True)
        bk.formula(ws, "E%d" % r, "=B%d*D%d" % (r, r), amt * w, fmt=MONEY, box=True)
        bk.formula(ws, "F%d" % r, "=B%d-E%d" % (r, r), amt * (1 - w), fmt=MONEY, box=True)
        bk.val(ws, "G%d" % r, b[5], box=True, align=WRAPTOP)
        ws.row_dimensions[r].height = 26
    t = last + 1
    bk.val(ws, "A%d" % t, "Total", font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "B%d" % t, "=SUM(B%d:B%d)" % (first, last), model["totalpv"],
               fmt=MONEY, font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "C%d" % t, "=SUM(C%d:C%d)" % (first, last), 1.0, fmt=PCT1, font=BOLD,
               box=True, fill=PALEFILL)
    bk.formula(ws, "D%d" % t, "=IFERROR(E%d/B%d, 0)" % (t, t),
               model["totalequity"] / model["totalpv"], fmt=PCT1, font=BOLD, box=True,
               fill=PALEFILL)
    bk.formula(ws, "E%d" % t, "=SUM(E%d:E%d)" % (first, last), model["totalequity"],
               fmt=MONEY, font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "F%d" % t, "=SUM(F%d:F%d)" % (first, last),
               model["totalpv"] - model["totalequity"], fmt=MONEY, font=BOLD, box=True,
               fill=PALEFILL)
    bk.val(ws, "G%d" % t, "Equity share of the whole plan", font=BOLD, box=True,
           fill=PALEFILL)
    bk.note(ws, t + 2, 7, NOTPROJECTION, height=28)
    ws.freeze_panes = "A4"
    return ws


def dashboard(bk, model):
    ws = bk.sheet("Summary Dashboard")
    bk.widths(ws, {"A": 36, "B": 20, "C": 14, "D": 24, "E": 28, "F": 14, "G": 22})
    bk.title(ws, 1, 7, "GOAL PLANNING SUMMARY")
    alloclast = 4 + len(BUCKETS)
    heads = [
        ("Total future value required:", "=Calculations!B3", model["totalfv"], BIGMONEY),
        ("Total cost increase over the plan:", "=Calculations!B4",
         model["totalfv"] - model["totalcost"], BIGMONEY),
        ("Total amount required today:", "=Calculations!B5", model["totalpv"], BIGMONEY),
        ("Number of goals:", "=Calculations!B6", len(GOALS), YEARS),
        ("Equity share of the plan:",
         "='Portfolio Allocation'!D%d" % alloclast,
         model["totalequity"] / model["totalpv"], PCT1),
    ]
    for i, (label, fml, cached, fmt) in enumerate(heads):
        r = 3 + i
        bk.val(ws, "A%d" % r, label, font=BOLD)
        bk.formula(ws, "B%d" % r, fml, cached, fmt=fmt, font=BOLD, box=True, fill=PALEFILL)

    bk.val(ws, "A9", "Goals registered", font=BOLD)
    bk.header(ws, 10, ["Goal Name", "Today's Cost (Rs)", "Horizon (yrs)",
                       "Inflation-Adjusted FV (Rs)", "Bucket", "Expected Return",
                       "Amount Required Today (Rs)"])
    for i in range(GOALROWS):
        r = 11 + i
        src = GOALROW0 + i
        g = model["goals"][i] if i < len(GOALS) else None
        cells = [
            ("A", "A", None, g["name"] if g else ""),
            ("B", "C", MONEY, g["cost"] if g else ""),
            ("C", "D", YEARS, g["years"] if g else ""),
            ("D", "G", MONEY, g["fv"] if g else ""),
            ("E", "I", None, g["bucket"] if g else ""),
            ("F", "J", PCT2, g["ret"] if g else ""),
            ("G", "K", MONEY, g["pv"] if g else ""),
        ]
        for outcol, srccol, fmt, cached in cells:
            bk.formula(ws, "%s%d" % (outcol, r),
                       "=IF('Goals Input'!%s%d=\"\",\"\",'Goals Input'!%s%d)"
                       % (srccol, src, srccol, src),
                       cached, fmt=fmt, box=True,
                       font=BOLD if outcol in ("A", "G") else BODY)
    t = 11 + GOALROWS
    bk.val(ws, "A%d" % t, "Total", font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "D%d" % t, "=SUM(D11:D%d)" % (t - 1), model["totalfv"], fmt=MONEY,
               font=BOLD, box=True, fill=PALEFILL)
    bk.formula(ws, "G%d" % t, "=SUM(G11:G%d)" % (t - 1), model["totalpv"], fmt=MONEY,
               font=BOLD, box=True, fill=PALEFILL)
    bk.note(ws, t + 2, 7, SERIESLIMIT, height=42)
    bk.note(ws, t + 3, 7, NOTPROJECTION, height=28,
            font=Font(name="Calibri", size=9, bold=True, color="C00000"))
    ws.freeze_panes = "A11"
    return ws


def corrections(bk, model):
    ws = bk.sheet("Corrections")
    bk.widths(ws, {"A": 26, "B": 46, "C": 46, "D": 62})
    bk.title(ws, 1, 4, "WHAT CHANGED FROM THE EARLIER VERSION")
    bk.note(ws, 2, 4,
            "Six corrections were made to the earlier version of this tool. Each one is "
            "recorded below with the reason for it.", height=20)
    bk.header(ws, 4, ["Item", "What the earlier version did", "What this version does",
                      "Why"])
    items = [
        ("Return assumptions",
         "18% for a 100% equity bucket, 15% for 75 equity to 25 debt, 12% for 50 to 50, "
         "10% for a short horizon 50 to 50, 8% for an 80% debt bucket and 6% for all debt.",
         "Two planning inputs, 11.0% for equity and 6.0% for debt, held in named cells on "
         "the Assumptions sheet. Every bucket return is built from those two.",
         "The NIFTY 50 compounded 13.02% on price over 20.8 years, but that window opens in "
         "April 2005 near the base of a bull run. From the January 2008 peak the same index "
         "compounded 8.27% on price over 18 years, and the 15 year rolling median is about "
         "10.8% on price. A measured dividend yield of about 1.27% takes that median to "
         "about 12.1% total return. 18% was never available to plan on."),
        ("Bucket returns",
         "Typed in one bucket at a time. The same 50 to 50 mix was priced at 12%, at 12% "
         "and at 10% in three different buckets, and the 80% debt bucket implied a 16% "
         "return on the 20% it held in hybrid funds.",
         "Each return is a formula on that bucket's own equity and debt weights and the two "
         "named inputs. The same mix now carries the same return everywhere it appears.",
         "A typed return drifts away from the mix it is supposed to describe, and this one "
         "already had. A formula cannot drift."),
        ("Priority mapping",
         "More equity went to higher priority goals. A high priority long horizon goal was "
         "sent to the 100% equity bucket.",
         "Less equity goes to higher priority goals. A high priority long horizon goal is "
         "sent to the 50 to 50 bucket, and the 100% equity bucket now holds low priority "
         "long horizon goals.",
         PRIORITYRULE + " The earlier mapping put the most risk behind the goals with the "
         "least room to fail, which is the wrong way round."),
        ("Equity inside two years",
         "Short horizon buckets held as much as 50% equity, and the short horizon 50 to 50 "
         "bucket was priced at 10%.",
         "No bucket inside two years holds equity. All three short horizon buckets earn the "
         "debt input and differ only in the duration of the debt they hold.",
         "Two years does not give an equity holding time to recover from a drawdown. A goal "
         "that lands inside one is a goal that is not funded."),
        ("The erosion column",
         "A column named Erosion (%) showed 3.83 for the retirement goal, rendered as "
         "382.8%. It was the nominal gross up ratio, future value less cost, divided by "
         "cost.",
         "That column is renamed Cost Increase (%), and a separate Purchasing Power Lost "
         "(%) column carries the real figure, one minus one divided by (1 plus inflation) "
         "raised to the horizon.",
         "Purchasing power lost is bounded at 100% by construction, so 382.8% was "
         "impossible on its face. For the retirement goal over 25 years at 6.5% the correct "
         "figure is 79.29%."),
        ("Limits stated on the sheet",
         "The tool carried no statement of what its numbers rested on or where they ran out.",
         "The Goals Input sheet, the Assumptions sheet and this sheet all carry the length "
         "of the evidence and the status of the numbers.",
         "The equity series is 20.9 years long and the retirement goal runs 25 years, so no "
         "25 year holding period can be measured from it. A client facing tool that shows "
         "projected returns without saying what they are is a regulatory exposure."),
    ]
    r = 5
    for item in items:
        for i, text in enumerate(item):
            bk.val(ws, "%s%d" % (get_column_letter(i + 1), r), text,
                   font=BOLD if i == 0 else BODY, align=WRAPTOP, box=True)
        ws.row_dimensions[r].height = max(
            46, 12 * math.ceil(max(len(item[1]) / 46.0, len(item[2]) / 46.0,
                                   len(item[3]) / 62.0)) + 10)
        r += 1

    bk.note(ws, r + 1, 4,
            "The sheet after this one prices all six corrections against the same five "
            "goals.", height=20)
    ws.freeze_panes = "A5"
    return ws


def impact(bk, model):
    ws = bk.sheet("Correction Impact")
    bk.title(ws, 1, 11, "WHAT THE CORRECTION IS WORTH")
    bk.note(ws, 2, 11,
            "Column E is the earlier version. Column H corrects the return inputs only and "
            "leaves the earlier priority mapping in place. Column K is this workbook, with "
            "the priority mapping corrected as well. All three discount the same five goals "
            "and the same inflation adjusted future values, so the difference between them "
            "is the correction and nothing else.", height=30)
    head = 4
    bk.header(ws, head, [
        "Goal", "Future value needed (Rs)", "Horizon (yrs)", "Earlier return",
        "Earlier amount today (Rs)", "Equity weight, earlier mapping",
        "Return on corrected inputs, earlier mapping", "Amount today on that basis (Rs)",
        "Bucket in this version", "Return in this version",
        "Amount today in this version (Rs)"], height=48)
    for c, w in zip("ABCDEFGHIJK",
                    [26, 17, 11, 12, 17, 15, 18, 18, 26, 13, 18]):
        ws.column_dimensions[c].width = w

    for i, g in enumerate(model["goals"]):
        rr = head + 1 + i
        src = GOALROW0 + i
        bk.formula(ws, "A%d" % rr, "='Goals Input'!A%d" % src, g["name"], font=BOLD,
                   box=True)
        bk.formula(ws, "B%d" % rr, "='Goals Input'!G%d" % src, g["fv"], fmt=MONEY, box=True)
        bk.formula(ws, "C%d" % rr, "='Goals Input'!D%d" % src, g["years"], fmt=YEARS,
                   box=True, align=Alignment(horizontal="center"))
        bk.val(ws, "D%d" % rr, PRIORRATE[i], fmt=PCT2, box=True)
        bk.formula(ws, "E%d" % rr, "=B%d/POWER(1+D%d,C%d)" % (rr, rr, rr),
                   g["priorpv"], fmt=MONEY, box=True)
        bk.val(ws, "F%d" % rr, PRIORWEIGHT[i], fmt=PCT1, box=True)
        bk.formula(ws, "G%d" % rr, "=F%d*EquityReturn+(1-F%d)*DebtReturn" % (rr, rr),
                   g["midret"], fmt=PCT2, box=True)
        bk.formula(ws, "H%d" % rr, "=B%d/POWER(1+G%d,C%d)" % (rr, rr, rr),
                   g["midpv"], fmt=MONEY, box=True)
        bk.formula(ws, "I%d" % rr, "='Goals Input'!I%d" % src, g["bucket"], box=True)
        bk.formula(ws, "J%d" % rr, "='Goals Input'!J%d" % src, g["ret"], fmt=PCT2, box=True)
        bk.formula(ws, "K%d" % rr, "='Goals Input'!K%d" % src, g["pv"], fmt=MONEY,
                   font=BOLD, box=True)
    tot = head + 1 + len(GOALS)
    bk.val(ws, "A%d" % tot, "Plan total", font=BOLD, box=True, fill=PALEFILL)
    for col, key in (("B", "totalfv"), ("E", "priortotal"), ("H", "midtotal"),
                     ("K", "totalpv")):
        bk.formula(ws, "%s%d" % (col, tot),
                   "=SUM(%s%d:%s%d)" % (col, head + 1, col, tot - 1), model[key],
                   fmt=MONEY, font=BOLD, box=True, fill=PALEFILL)

    retrow = head + 1 + 4
    lines = [
        ('="The retirement goal needed Rs "&TEXT(E{r},"#,##0")&" under the earlier '
         'version, Rs "&TEXT(H{r},"#,##0")&" once the return inputs are corrected, and '
         'Rs "&TEXT(K{r},"#,##0")&" in this workbook."'.format(r=retrow),
         "The retirement goal needed Rs %s under the earlier version, Rs %s once the return "
         "inputs are corrected, and Rs %s in this workbook."
         % (fmtnum(model["goals"][4]["priorpv"]), fmtnum(model["goals"][4]["midpv"]),
            fmtnum(model["goals"][4]["pv"]))),
        ('="The five goal plan needed Rs "&TEXT(E{t},"#,##0")&" under the earlier version, '
         'Rs "&TEXT(H{t},"#,##0")&" once the return inputs are corrected, and Rs "&'
         'TEXT(K{t},"#,##0")&" in this workbook."'.format(t=tot),
         "The five goal plan needed Rs %s under the earlier version, Rs %s once the return "
         "inputs are corrected, and Rs %s in this workbook."
         % (fmtnum(model["priortotal"]), fmtnum(model["midtotal"]),
            fmtnum(model["totalpv"]))),
        ('="Two figures were quoted for this rebuild before it was run, Rs 1.42 crore for '
         'retirement and Rs 4.12 crore for the plan. Those are the column H figures, which '
         'correct the return inputs and leave the priority mapping as it was: this workbook '
         'reproduces them at Rs "&TEXT(H{r}/10000000,"0.00")&" crore and Rs "&'
         'TEXT(H{t}/10000000,"0.00")&" crore. Correcting the priority mapping as well moves '
         'them to Rs "&TEXT(K{r}/10000000,"0.00")&" crore and Rs "&'
         'TEXT(K{t}/10000000,"0.00")&" crore, because the two long horizon high priority '
         'goals now discount at the 50 to 50 rate rather than at the all equity rate."'
         .format(r=retrow, t=tot),
         "Two figures were quoted for this rebuild before it was run, Rs 1.42 crore for "
         "retirement and Rs 4.12 crore for the plan. Those are the column H figures, which "
         "correct the return inputs and leave the priority mapping as it was: this workbook "
         "reproduces them at Rs %0.2f crore and Rs %0.2f crore. Correcting the priority "
         "mapping as well moves them to Rs %0.2f crore and Rs %0.2f crore, because the two "
         "long horizon high priority goals now discount at the 50 to 50 rate rather than at "
         "the all equity rate."
         % (model["goals"][4]["midpv"] / 1e7, model["midtotal"] / 1e7,
            model["goals"][4]["pv"] / 1e7, model["totalpv"] / 1e7)),
    ]
    rr = tot + 2
    for fml, cached in lines:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=11)
        bk.formula(ws, "A%d" % rr, fml, cached, font=BOLD, align=WRAPTOP)
        ws.row_dimensions[rr].height = max(30, 14 * math.ceil(len(cached) / 175.0) + 6)
        rr += 1
    bk.note(ws, rr + 1, 11, NOTPROJECTION, height=28,
            font=Font(name="Calibri", size=9, bold=True, color="C00000"))
    ws.freeze_panes = "A5"
    return ws


def fmtnum(x):
    return "{:,.0f}".format(x)


def buildmodel():
    goals = []
    for i, (name, cat, cost, years, priority) in enumerate(GOALS):
        infl = MIDRATE[cat]
        fv = cost * (1.0 + infl) ** years
        bucket = bucketfor(years, priority)
        ret = bucketreturn(bucket)
        pv = fv / (1.0 + ret) ** years
        midret = PRIORWEIGHT[i] * EQUITY + (1.0 - PRIORWEIGHT[i]) * DEBT
        goals.append({
            "name": name, "category": cat, "cost": cost, "years": years,
            "priority": priority, "infl": infl, "fv": fv, "bucket": bucket, "ret": ret,
            "pv": pv, "costpct": (fv - cost) / cost,
            "lost": 1.0 - 1.0 / (1.0 + infl) ** years,
            "priorpv": fv / (1.0 + PRIORRATE[i]) ** years,
            "midret": midret, "midpv": fv / (1.0 + midret) ** years,
        })
    model = {
        "goals": goals,
        "totalcost": sum(g["cost"] for g in goals),
        "totalfv": sum(g["fv"] for g in goals),
        "totalpv": sum(g["pv"] for g in goals),
        "priortotal": sum(g["priorpv"] for g in goals),
        "midtotal": sum(g["midpv"] for g in goals),
    }
    model["bucketpv"] = {b[1]: sum(g["pv"] for g in goals if g["bucket"] == b[1])
                         for b in BUCKETS}
    model["bucketcount"] = {b[1]: sum(1 for g in goals if g["bucket"] == b[1])
                            for b in BUCKETS}
    model["totalequity"] = sum(g["pv"] * EQUITYWEIGHT[g["bucket"]] for g in goals)
    return model


def build():
    model = buildmodel()
    bk = Book()
    goalsinput(bk, model)
    inflationsheet(bk)
    assumptions(bk, model)
    calculations(bk, model)
    allocation(bk, model)
    dashboard(bk, model)
    corrections(bk, model)
    impact(bk, model)

    bk.wb.defined_names.add(DefinedName("EquityReturn", attr_text="Assumptions!$B$2"))
    bk.wb.defined_names.add(DefinedName("DebtReturn", attr_text="Assumptions!$D$2"))
    bk.wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    bk.wb.save(OUT)
    n = injectcached(OUT, bk.cache)
    return model, n


def report(model, injected):
    print("Written: %s" % OUT)
    print("Formula cells given a calculated value: %d" % injected)
    print("")
    print("Planning inputs: equity %0.2f%%, debt %0.2f%%" % (EQUITY * 100, DEBT * 100))
    print("")
    print("%-26s %5s %-30s %7s %16s" % ("Goal", "Yrs", "Bucket", "Return", "Needed today"))
    for g in model["goals"]:
        print("%-26s %5d %-30s %6.2f%% %16s"
              % (g["name"], g["years"], g["bucket"], g["ret"] * 100, fmtnum(g["pv"])))
    print("%-26s %5s %-30s %7s %16s"
          % ("Plan total", "", "", "", fmtnum(model["totalpv"])))
    print("")
    print("Retirement needed today: Rs %s (Rs %0.2f crore)"
          % (fmtnum(model["goals"][4]["pv"]), model["goals"][4]["pv"] / 1e7))
    print("Plan total needed today: Rs %s (Rs %0.2f crore)"
          % (fmtnum(model["totalpv"]), model["totalpv"] / 1e7))


if __name__ == "__main__":
    m, injected = build()
    report(m, injected)
