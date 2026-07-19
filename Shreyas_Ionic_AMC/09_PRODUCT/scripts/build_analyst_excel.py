"""
FROZEN internal/analyst full-detail Excel for STOCK_SCORECARD_750 holdings reviews.
Implements FROZEN_METHODOLOGY.md 'Output schema — internal/analyst (full detail)':
identity | cyclicality | value | all pillar sub-scores (both horizons) | composites |
gates/penalty/boost | final adj scores | quant recs | full qualitative fields |
technical-agent fields (blank until that pass runs on a batch).
Rerunnable: python build_analyst_excel.py <quant_csv> <qual_json_dir> <out_xlsx>
"""
import os, sys, json, glob
os.environ["PYTHONIOENCODING"] = "utf-8"
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ionic Wealth house style (00_GOVERNANCE/STYLE_GUIDE.md) — shared module
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ionic_style import (
    INK, NAVY, GOLD, TEAL, RUST, STONE, detell,
    BODY_FONT_NAME, HEAD_FONT_NAME,
    ROW_BORDER as BORDER, HEADER_BORDER,
)
FONT = BODY_FONT_NAME  # Georgia body, per house style
# header banding by column section (identity/pillars/gates+finals/analyst/longtext/technical)
SECTION_FILLS = {
    "identity": PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid"),
    "pillars": PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid"),
    "finals": PatternFill(start_color=STONE, end_color=STONE, fill_type="solid"),
    "analyst": PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid"),
    "text": PatternFill(start_color="30506B", end_color="30506B", fill_type="solid"),
    "technical": PatternFill(start_color="707070", end_color="707070", fill_type="solid"),
}
def _section(j):  # 1-based column index -> section key (matches COLUMNS layout)
    if j <= 7: return "identity"
    if j <= 20: return "pillars"
    if j <= 32: return "finals"
    if j <= 36: return "analyst"
    if j <= 43: return "text"
    return "technical"
HEADER_FONT = Font(name=HEAD_FONT_NAME, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name=FONT, size=10, color=INK)
SELL_FILL = PatternFill(start_color="F1DFD7", end_color="F1DFD7", fill_type="solid")
SELL_FONT = Font(name=FONT, size=10, bold=True, color=RUST)
HOLD_FILL = PatternFill(start_color="DFE9E6", end_color="DFE9E6", fill_type="solid")
HOLD_FONT = Font(name=FONT, size=10, bold=True, color=TEAL)
ESC_FILL = PatternFill(start_color="F3EAD9", end_color="F3EAD9", fill_type="solid")

# (header, quant-CSV column or QUAL:<field>, kind) — kind: num1 = 1dp number, num0 = int,
# pct = 2dp, text = plain, wrap = wrapped long text
COLUMNS = [
    ("Symbol", "symbol", "text"),
    ("Company Name", "company_name", "text"),
    ("ISIN", "isin", "text"),
    ("Sector", "sector", "text"),
    ("Cyclicality", "cyclicality_tag", "text"),
    ("Value (Rs)", "value_inr", "num0"),
    ("% of Portfolio", "PCT", "pct"),
    ("Quality", "quality_score", "num1"),
    ("Growth 3Y", "growth_3y_score", "num1"),
    ("Growth 1Y", "growth_1y_score", "num1"),
    ("Value Pillar", "value_score", "num1"),
    ("Stage 3Y", "stage_3y_score", "num1"),
    ("Stage 1Y", "stage_1y_score", "num1"),
    ("Stage Timing", "stage_timing_tag", "text"),
    ("SectorMacro 3Y", "sector_macro_3y_score", "num1"),
    ("SectorMacro 1Y", "sector_macro_1y_score", "num1"),
    ("OwnFlow 3Y", "ownership_flow_3y_score", "num1"),
    ("OwnFlow 1Y", "ownership_flow_1y_score", "num1"),
    ("Accum 3Y", "accumulation_3y_score", "num1"),
    ("Accum 1Y", "accumulation_1y_score", "num1"),
    ("Composite 3Y", "composite_3y", "num1"),
    ("Composite 1Y", "composite_1y", "num1"),
    ("BS Flag", "bs_flag", "text"),
    ("Liq Flag", "liquidity_flag", "text"),
    ("Red Flags", "redflag_count", "num0"),
    ("Penalty", "penalty", "num0"),
    ("Boost", "boost", "num0"),
    ("Final 3Y (adj)", "final_3y_adj", "num1"),
    ("Final 1Y (adj)", "final_1y_adj", "num1"),
    ("Quant Rec 3Y", "recommendation_3y", "text"),
    ("Quant Rec 1Y", "recommendation_1y", "text"),
    ("Quant Rec Overall", "recommendation", "text"),
    ("ANALYST REC (FINAL)", "QUAL:your_recommendation", "text"),
    ("Est. 3Y Growth (%/yr)", "QUAL:expected_next_3y_growth_pct", "num1"),
    ("Escalation", "QUAL:escalation_flag", "text"),
    ("Escalation Reason", "QUAL:escalation_reason", "wrap"),
    ("Summary", "QUAL:summary", "wrap"),
    ("Detailed Rationale", "QUAL:detailed_rationale", "wrap"),
    ("Recommendation Rationale", "QUAL:recommendation_rationale", "wrap"),
    ("Positives", "QUAL:positive_para", "wrap"),
    ("Negatives", "QUAL:negative_para", "wrap"),
    ("Reverse-DCF Judgment", "QUAL:reverse_dcf_judgment", "wrap"),
    ("Research Sources", "QUAL:research_sources", "wrap"),
    ("Chart LT Technical Score", "QUAL:chart_long_term_technical_pattern_score", "num1"),
    ("Choppiness Flag", "QUAL:choppiness_flag", "text"),
    ("Technical Commentary", "QUAL:technical_commentary", "wrap"),
]


def build(quant_csv: str, qual_json_dir: str, out_xlsx: str):
    quant = pd.read_csv(quant_csv)
    sym_col = "Symbol" if "Symbol" in quant.columns else "symbol"
    total_val = quant["value_inr"].sum() if "value_inr" in quant.columns else None

    quals = {}
    for pat in ("pf_qual_*.json", "qual_*.json"):
        for p in glob.glob(os.path.join(qual_json_dir, pat)):
            d = json.load(open(p, encoding="utf-8"))
            quals.setdefault(d["symbol"], d)

    recs = []
    for _, r in quant.iterrows():
        sym = r[sym_col]
        q = quals.get(sym, {})
        row = {}
        for hdr, src, kind in COLUMNS:
            if src == "PCT":
                v = (r["value_inr"] / total_val * 100) if total_val else None
            elif src.startswith("QUAL:"):
                v = q.get(src[5:], None)
                if isinstance(v, list):
                    v = "\n".join(str(x) for x in v)
                if isinstance(v, bool):
                    v = "ESCALATED" if v else ""
                if isinstance(v, str) and kind == "wrap":
                    v = detell(v)  # house de-AI-ification pass on all research prose
            else:
                v = r.get(src, None)
            row[hdr] = v
        recs.append(row)
    df = pd.DataFrame(recs)

    # Technical-agent columns render ONLY when that pass has run (Principal order
    # 2026-07-19: no dead columns). The technical pass stays in the methodology as
    # an optional timing overlay; when any value exists the columns come back.
    TECH_COLS = ["Chart LT Technical Score", "Choppiness Flag", "Technical Commentary"]
    def _empty(v):
        return v is None or (isinstance(v, float) and pd.isna(v)) or (isinstance(v, str) and not v.strip())
    tech_ran = not all(df[c].map(_empty).all() for c in TECH_COLS if c in df.columns)
    if not tech_ran:
        df = df.drop(columns=[c for c in TECH_COLS if c in df.columns])

    # same ordering as the client sheet: Sell-first (final analyst rec), largest first
    rec_order = {"Sell": 0, "Hold": 1, "No Recommendation": 2}
    df["_rk"] = df["ANALYST REC (FINAL)"].map(rec_order).fillna(3)
    df = df.sort_values(["_rk", "Value (Rs)"], ascending=[True, False]).drop(columns="_rk")

    wb = Workbook()
    ws = wb.active
    ws.title = "Analyst Full Detail"
    kinds = {h: k for h, _, k in COLUMNS}
    for j, h in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill, c.border = HEADER_FONT, SECTION_FILLS[_section(j)], BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    rec_col = list(df.columns).index("ANALYST REC (FINAL)") + 1
    esc_col = list(df.columns).index("Escalation") + 1
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        for j, h in enumerate(df.columns, start=1):
            v = r[h]
            k = kinds[h]
            if pd.isna(v) if not isinstance(v, (str, list)) else False:
                v = None
            if v is not None and k in ("num1", "pct"):
                try:
                    v = round(float(v), 2 if k == "pct" else 1)
                except (TypeError, ValueError):
                    pass
            if v is not None and k == "num0":
                try:
                    v = int(round(float(v)))
                except (TypeError, ValueError):
                    pass
            c = ws.cell(row=i, column=j, value=v)
            c.font, c.border = BODY_FONT, BORDER
            if k in ("num1", "num0", "pct"):
                c.alignment = Alignment(vertical="top", horizontal="right")
                c.number_format = {"pct": "0.00", "num1": "0.0", "num0": "0"}[k]
            else:
                c.alignment = Alignment(vertical="top", wrap_text=(k == "wrap"))
        rc = ws.cell(row=i, column=rec_col)
        if r["ANALYST REC (FINAL)"] == "Sell":
            rc.font, rc.fill = SELL_FONT, SELL_FILL
        elif r["ANALYST REC (FINAL)"] == "Hold":
            rc.font, rc.fill = HOLD_FONT, HOLD_FILL
        if r["Escalation"] == "ESCALATED":
            ws.cell(row=i, column=esc_col).fill = ESC_FILL
        ws.row_dimensions[i].height = 110

    for j, h in enumerate(df.columns, start=1):
        k = kinds[h]
        ws.column_dimensions[get_column_letter(j)].width = (
            60 if k == "wrap" else 22 if h in ("Company Name", "Sector") else
            15 if h in ("ISIN", "ANALYST REC (FINAL)", "Quant Rec Overall") else 11)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # collapsible column groups: pillars (H-T), gates/finals (U-AF), long text (AK-AQ), technical (AR-AT)
    groups = [("H", "T"), ("U", "AF"), ("AK", "AQ")] + ([("AR", "AT")] if tech_ran else [])
    for a_, b_ in groups:
        ws.column_dimensions.group(a_, b_, outline_level=1, hidden=False)

    # score heat-map on numeric score columns
    from openpyxl.formatting.rule import ColorScaleRule
    last = len(df) + 1
    scale = lambda: ColorScaleRule(start_type="num", start_value=20, start_color="E8B4B4",
                                   mid_type="num", mid_value=50, mid_color="F5E6B8",
                                   end_type="num", end_value=80, end_color="B8D8C7")
    for col in ("H", "I", "J", "K", "L", "M", "O", "P", "Q", "R", "S", "T", "U", "V", "AB", "AC"):
        ws.conditional_formatting.add(f"{col}2:{col}{last}", scale())

    _field_guide_sheet(wb)
    _reader_sheet(wb, df)
    extra = _pf_analytics_sheet(wb, qual_json_dir)

    wb.properties.creator = "Ionic Wealth"
    wb.properties.title = "STOCK_SCORECARD_750 - Analyst Full Detail"
    wb.save(out_xlsx)
    print(f"Saved {len(df)} rows x {len(df.columns)} cols to {out_xlsx} "
          f"(+ Field Guide + Research Reader{' + Portfolio Analytics (Full)' if extra else ''})")
    return out_xlsx


def _pf_analytics_sheet(wb, qual_json_dir):
    """Internal full analytics: everything the client sheet shows PLUS the factor
    regression, forward-alpha [ESTIMATE] and full assumptions. Portfolio-holdings
    reviews only; skipped when pf_analytics.json is absent (750-universe mode)."""
    p = os.path.join(qual_json_dir, "pf_analytics.json")
    if not os.path.exists(p):
        return False
    an = json.load(open(p, encoding="utf-8"))
    ws = wb.create_sheet("Portfolio Analytics (Full)")
    ws.sheet_properties.tabColor = TEAL
    TITLE = Font(name=HEAD_FONT_NAME, size=13, bold=True, color=NAVY)
    SECT = Font(name=HEAD_FONT_NAME, size=11, bold=True, color=NAVY)
    KEY = Font(name=FONT, size=10, bold=True, color=INK)
    VAL = Font(name=FONT, size=10, color=INK)
    SUB = Font(name=FONT, size=9, italic=True, color=STONE)

    r = 1
    ws.cell(row=r, column=1, value="Portfolio Analytics (Full) · internal").font = TITLE
    r += 1
    ws.cell(row=r, column=1, value=f"As of {an.get('as_of', '')} · engine: compute_portfolio_analytics.py "
            "· client-facing subset lives in the Ionic Wealth workbook").font = SUB
    r += 2

    def kv_block(r, title, pairs):
        ws.cell(row=r, column=1, value=title).font = SECT
        r += 1
        for k, v in pairs:
            ws.cell(row=r, column=1, value=k).font = KEY
            ws.cell(row=r, column=1).border = BORDER
            c = ws.cell(row=r, column=2, value=v)
            c.font, c.border = VAL, BORDER
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if isinstance(v, str) and len(v) > 90:
                ws.row_dimensions[r].height = 13 * (len(v) // 90 + 1) + 6
            r += 1
        return r + 1

    for label, key in [("Return and risk — 3 years (simulated, vs Nifty 50 TRI proxy)", "portfolio_3y"),
                       ("Return and risk — 1 year", "portfolio_1y")]:
        d = an.get(key, {})
        b = an.get(key.replace("portfolio", "bench"), {})
        pairs = [(m.replace("_", " "), f"{d.get(m)}   (Nifty 50: {b.get(m)})" if m in b else d.get(m))
                 for m in ["cagr_pct", "vol_pct", "sharpe", "sortino", "max_dd_pct", "beta",
                           "alpha_ann_pct", "tracking_error_pct", "info_ratio",
                           "up_capture", "down_capture", "corr_bench"] if m in d]
        r = kv_block(r, detell(label), pairs)

    fac = an.get("factors_3y", {})
    if fac:
        pairs = [("R-squared", fac.get("r2")), ("n (days)", fac.get("n_days")),
                 ("alpha (annualized %)", f"{fac.get('alpha_ann_pct')}  (t = {fac.get('alpha_t')})")]
        for f in ("MKT", "SIZE", "VALUE", "MOM"):
            d = fac.get(f, {})
            pairs.append((f + " beta", f"{d.get('beta')}  (t = {d.get('t')})"))
        r = kv_block(r, "Factor regression, daily 3y (MKT excess / SIZE Mid-N50 / VALUE V50-N500 / MOM M30-N500)", pairs)

    fv = an.get("forward_view", {})
    if fv:
        r = kv_block(r, "Forward view (analyst layer, weighted)", [
            ("Weighted forward growth (%)", fv.get("weighted_fwd_growth_pct")),
            ("Weight with growth >= 15% (%)", fv.get("weight_growth_ge_15_pct")),
            ("Weight with growth < 10% (%)", fv.get("weight_growth_lt_10_pct")),
            ("Nifty 50 trailing EPS CAGR 3y (%)", fv.get("nifty50_trailing_eps_cagr_3y_pct")),
            ("Expected-alpha note", detell(fv.get("expected_alpha_note", ""))),
        ])

    vc = an.get("valuation_context", {})
    if vc:
        pairs = [(k, f"P/E {d.get('pe_now')} vs median {d.get('pe_median')} "
                  f"({d.get('pe_pctile_since2016'):.0f}th pctile since 2016)") for k, d in vc.items()]
        r = kv_block(r, "Index valuation context", pairs)

    asm = an.get("assumptions", {})
    r = kv_block(r, "Assumptions", [(k.replace("_", " "), str(v)) for k, v in asm.items()])
    ws.cell(row=r, column=1, value="Top-15 correlation matrix: pf_corr_matrix.csv (rendered as a heatmap "
            "in the client workbook's Portfolio Analytics sheet).").font = SUB

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 95
    ws.sheet_view.showGridLines = False
    return True


FIELD_DEFS = [
    ("Symbol / Company / ISIN / Sector / Cyclicality", "Identity from the quant universe; cyclicality tag drives the regime weight tilt."),
    ("Value (Rs) / % of Portfolio", "Position value from client CAS; blank in 750-universe mode."),
    ("Quality", "mean(sector-neutral pctile ROE, ROCE). Financials: judge CRAR/asset quality instead of D/E."),
    ("Growth 3Y / 1Y", "Percentile of 3yr revenue CAGR / TTM revenue growth (TRAILING - the forward view lives in the analyst's growth estimate)."),
    ("Value Pillar", "Blend of -PE (universe + sector-tier), -PB, FCF-yield percentiles."),
    ("Stage 3Y / 1Y / Timing", "Return percentiles gated by 200/50DMA; replaced by the technical agent's chart score once that pass runs."),
    ("SectorMacro / OwnFlow / Accum", "Sector-mean return + regime fit; FII+DII flow percentile; OBV-slope percentile."),
    ("Composite 3Y / 1Y", "Weighted pillar blend (63/37 fundamentals-tilt 3Y; 40/60 technical-tilt 1Y)."),
    ("BS Flag / Liq Flag", "Balance-sheet gate (RED caps score at 40; AMBER x0.85; financials exempt from D/E trigger -> 'N/A-financial-sector') / turnover gate."),
    ("Red Flags / Penalty / Boost", "Flags: IntCov<1.5; D/E>2.5 (non-financials); negative 1y revenue; >15pp deceleration; analyst forward growth <10%. Penalty = -(2^n - 1) capped -10. Boost +3 = zero flags + Quality&Value >60pctile."),
    ("Final 3Y/1Y (adj)", "Composite after gates + penalty + boost. These feed the client Ionic Score (0.6/0.4 blend + forward adjustment)."),
    ("Quant Rec 3Y/1Y/Overall", ">=40 Hold, <40 Sell per horizon; overall Sell if either horizon Sell. The ANALYST REC overrides all of it."),
    ("ANALYST REC (FINAL)", "The governing call. Sell or Hold only (client layer may convert to Trim on concentration)."),
    ("Est. 3Y Growth (%/yr)", "Analyst's FORWARD 3-5y growth view (never trailing). Client-score impact (v6.2): <5% -15 | 5-10% -5 | 10-15% 0 | 15-20% +5 | 20-25% +10 | >=25% +15 | +20 exceptional (>=25% + ROE>=20% + low dilution)."),
    ("Escalation / Reason", "Narrow: genuine Hold-vs-Sell coin-flips or methodology gaps only. Staleness is never escalated. Principal adjudicates."),
    ("Summary...Research Sources", "The qualitative record. Summary = the exact client-sheet rationale. Full texts readable in the Research Reader sheet."),
    ("Chart LT Technical Score / Choppiness / Commentary", "Separate technical-agent pass (monthly-chart judgment). Columns appear only when that pass has run on the batch (2026-07-19: hidden when empty)."),
]


def _field_guide_sheet(wb):
    ws = wb.create_sheet("Field Guide")
    ws.sheet_properties.tabColor = "5F5E57"
    c = ws.cell(row=1, column=1, value="Field Guide · STOCK_SCORECARD_750 analyst workbook (methodology: FROZEN_METHODOLOGY.md v6.2, 2026-07-18)")
    c.font = Font(name=FONT, size=12, bold=True, color="1F3864")
    r = 3
    for j, h in enumerate(("Column(s)", "Definition / frozen rule"), 1):
        cc = ws.cell(row=r, column=j, value=h)
        cc.font, cc.fill, cc.border = HEADER_FONT, HEADER_FILL, BORDER
    r += 1
    for name, desc in FIELD_DEFS:
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
        d = ws.cell(row=r, column=2, value=detell(desc))
        d.font = BODY_FONT
        d.alignment = Alignment(wrap_text=True, vertical="top")
        for j in (1, 2):
            ws.cell(row=r, column=j).border = BORDER
        ws.row_dimensions[r].height = max(28, 13 * (len(desc) // 95 + 1) + 6)
        r += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 105
    ws.sheet_view.showGridLines = False


def _chunks(text, size=1100):
    """Split long text at sentence-ish boundaries so every chunk fits a readable row."""
    if not isinstance(text, str) or not text.strip():
        return []
    text = text.strip()
    out, cur = [], ""
    for part in text.replace("\n", " \n").split(". "):
        cand = (cur + ". " + part) if cur else part
        if len(cand) > size and cur:
            out.append(cur.strip() + ".")
            cur = part
        else:
            cur = cand
    if cur.strip():
        out.append(cur.strip())
    return out


READER_FIELDS = [("Summary", "Summary"), ("Recommendation rationale", "Recommendation Rationale"),
                 ("Positives", "Positives"), ("Negatives", "Negatives"),
                 ("Reverse-DCF judgment", "Reverse-DCF Judgment"),
                 ("Detailed rationale", "Detailed Rationale"),
                 ("Escalation reason", "Escalation Reason"), ("Research sources", "Research Sources")]


def _reader_sheet(wb, df):
    """Per-stock readable blocks - NOTHING clipped (long fields chunked across rows)."""
    ws = wb.create_sheet("Research Reader")
    ws.sheet_properties.tabColor = "B08D57"
    c = ws.cell(row=1, column=1, value="Research Reader · full analyst text per stock (Ctrl+F a ticker to jump). Same order as the main sheet.")
    c.font = Font(name=FONT, size=11, bold=True, color="1F3864")
    r = 3
    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for _, row in df.iterrows():
        head = (f'{row["Symbol"]}  -  {row["Company Name"]}   |   FINAL: {row["ANALYST REC (FINAL)"]}'
                f'   |   Final 3Y/1Y: {row["Final 3Y (adj)"]}/{row["Final 1Y (adj)"]}'
                f'   |   Fwd growth: {row["Est. 3Y Growth (%/yr)"]}%'
                + ("   |   ESCALATED" if row["Escalation"] == "ESCALATED" else ""))
        hc = ws.cell(row=r, column=1, value=head)
        hc.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
        hc.fill = hdr_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
        for label, col in READER_FIELDS:
            for k, chunk in enumerate(_chunks(row[col])):
                ws.cell(row=r, column=1, value=label if k == 0 else "").font = Font(name=FONT, size=9.5, bold=True, color="5F5E57")
                t = ws.cell(row=r, column=2, value=chunk)
                t.font = Font(name=FONT, size=10)
                t.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r].height = min(400, 13.5 * (len(chunk) // 105 + 1) + 6)
                r += 1
        r += 1  # spacer between stocks
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 115
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    base = r"C:\Users\Shreyas.1Gupta\OneDrive - Angel Broking Limited\Desktop\Backup\NIFTY 500\Shreyas_Ionic_AMC"
    quant_csv = sys.argv[1] if len(sys.argv) > 1 else base + r"\04_RND_LAB\STOCK_SCORECARD_750\results\portfolio_quant.csv"
    qual_dir = sys.argv[2] if len(sys.argv) > 2 else base + r"\04_RND_LAB\STOCK_SCORECARD_750\results"
    out = sys.argv[3] if len(sys.argv) > 3 else base + r"\09_PRODUCT\reports\ANALYST_RECOMMENDATIONS.xlsx"
    build(quant_csv, qual_dir, out)
