# -*- coding: utf-8 -*-
"""build_scores_excel.py — the 750-universe scorecard Excel, now carrying the FIVE SIGNALS.
v8, 2026-08-07 (Principal): "in excel we had of 750 stock we have to add that while stock research
itself to avoid confusion and recommendation forward data".

What changed from the 2026-07-21 version:
  * The five client-deck signals (Quality / Growth / Value / Technical / Sector & Flows) appear as
    colour-banded word cells, computed by the SAME `pr_template/lib/five_signals.py` the deck uses --
    one source of truth, so research and client pages can never disagree about what green means.
  * Forward data: the analyst's expected 3-5y growth (N100 research run + portfolio pf_qual files)
    joins in; where present it blends into the Growth signal exactly as on the deck page.
  * v3 corrected scores sit BESIDE v1 (thin-history fix, results/THIN_COVERAGE_FIX_NOTE.md):
    1y-sibling substitution, neutral-50 fill, growth-artefact neutralisation, March-to-March growth.
  * New flags: Thin history / growth artefact / one-time-income risk / PAT-Sales divergence.

Reads results/full750_scored_v3.csv for the corrected scores and results/full750_scored.csv for the
v1 comparison column. v1 is the ENGINE OUTPUT and the input the v3 corrector reads -- it is not a
superseded duplicate, and deleting it breaks the whole chain.
Internal research tool; not investment advice.
"""
import glob
import json
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(HERE, "..", "pr_template", "lib")))
import five_signals as F                                                   # noqa: E402


def _nifty_root(p):
    while True:
        p, tail = os.path.split(p)
        if not tail:
            raise RuntimeError("NIFTY 500 root not found")
        cand = os.path.join(p, tail)
        if os.path.isdir(os.path.join(cand, "Shreyas_Ionic_AMC")) or tail == "NIFTY 500":
            return cand


ROOT = _nifty_root(HERE)
RES = os.path.join(ROOT, "Shreyas_Ionic_AMC", "04_RND_LAB", "STOCK_SCORECARD_750", "results")
SRC_V3 = os.path.join(RES, "full750_scored_v3.csv")
SRC_V1 = os.path.join(RES, "full750_scored.csv")
OUT = os.path.join(ROOT, "Shreyas_Ionic_AMC", "09_PRODUCT", "reports",
                   "NIFTY750_SCORECARD_20260807.xlsx")

AS_OF = "2026-08-07"

# ---- forward growth estimates: every research output that carries one ------------------------------
def load_forward():
    fwd = {}
    n100 = os.path.join(RES, "N100_RESEARCH_SUMMARY.csv")
    if os.path.exists(n100):
        t = pd.read_csv(n100)
        for _, r in t.iterrows():
            v = pd.to_numeric(r.get("expected_next_3y_growth_pct"), errors="coerce")
            if pd.notna(v):
                fwd[str(r["symbol"]).strip().upper()] = float(v)
    for p in glob.glob(os.path.join(RES, "pf_qual_*.json")):
        sym = os.path.basename(p)[len("pf_qual_"):-len(".json")].upper()
        try:
            with open(p, "r", encoding="utf-8") as fh:
                v = json.load(fh).get("expected_next_3y_growth_pct")
            if v is not None:
                fwd[sym] = float(v)          # per-name research beats the batch summary
        except (OSError, ValueError, TypeError):
            continue
    return fwd


src = SRC_V3 if os.path.exists(SRC_V3) else SRC_V1
df = pd.read_csv(src)
df["roe"] = df["roe"] * 100
df["roce"] = df["roce"] * 100
FWD = load_forward()
df["fwd_eps_growth_pct"] = df["symbol"].astype(str).str.upper().map(FWD)
has_v3 = "final_score_3y_v3" in df.columns

# Five signals per row, via the deck's own lib. The forward EPS estimate is NOT blended into Growth --
# Growth is a trailing revenue-CAGR percentile and the estimate is expected EPS growth (Principal,
# 2026-08-07). It appears as its own column instead, where it reads as what it is.
# Signals read the *_v3 (imputed) pillars where v3 produced them, so a dot never says "not scored" on a
# name whose score used a substituted value. `growth_pct` carries the analyst's expected EPS growth into
# the Growth signal, which weights it 60 against 40 on trailing revenue (Principal, 2026-08-07).
_PILL_V3 = [c[:-3] for c in df.columns if c.endswith("_v3") and c[:-3].endswith("_score")]


def _sig_row(r):
    rec = dict(r)
    for p in _PILL_V3:
        v = r.get(f"{p}_v3")
        if pd.notna(v):
            rec[p] = v
    if pd.notna(r.get("fwd_eps_growth_pct")):
        rec["growth_pct"] = float(r["fwd_eps_growth_pct"])
    return {c: v for c, v in F.signals(rec)}


_sig_rows = [_sig_row(r) for _, r in df.iterrows()]
for cat in F.CATS:
    df[f"sig::{cat}"] = [F.word(sr[cat]) for sr in _sig_rows]

sort_col = "ionic_score_v3" if has_v3 else "final_score_3y"
df = df.sort_values(sort_col, ascending=False, na_position="last").reset_index(drop=True)

COLS = [("symbol", "Symbol", None), ("sector", "Sector", None)]
if has_v3:
    # ONE column named "Call". The v1 call is deliberately NOT shown beside it: two adjacent columns
    # both labelled "Call" and giving opposite answers is a trap, and it caught the Principal twice --
    # HOMEFIRST reads Hold on v3 (64.1) and Sell on v1 (39.6), and the eye lands on whichever is
    # nearer. v3 is frozen, so it is the call; the v1 SCORE stays for comparison, the v1 verdict does
    # not. 62 names disagree between the two, so this is not a cosmetic choice.
    COLS += [("recommendation_v3", "Call", None), ("ionic_score_v3", "Ionic Score", 1)]
else:
    COLS += [("recommendation_overall", "Call", None)]
COLS += [("final_score_3y", "Score 3Y v1", 1)]
if has_v3:
    COLS += [("final_score_3y_v3", "Score 3Y v3", 1)]
COLS += [("final_score_1y", "Score 1Y v1", 1)]
if has_v3:
    COLS += [("final_score_1y_v3", "Score 1Y v3", 1)]
COLS += [(f"sig::{c}", c, None) for c in F.CATS]
COLS += [("fwd_eps_growth_pct", "Fwd EPS Grw % (analyst)", 0)]
if has_v3:
    COLS += [("trim_eligible_v3", "Trim eligible (why)", None),
             ("analyst_call", "Analyst call", None), ("base_score_v3", "Base score", 1),
             ("fwd_growth_input_pct", "Fwd grw used %", 1),
             ("fwd_growth_points", "Grw pts", 0), ("conviction_points", "Conv pts", 0),
             ("forward_adjustment", "Fwd adj", 0),
             ("analyst_conversion", "Analyst conversion", None)]
    COLS += [("history_class", "History", None), ("pillars_observed", "Pillars /7", 0),
             ("imputation_applied", "Imputation used", None),
             ("listing_return_pctile", "Listing-ret pctile", 0),
             ("growth_artifact_flag", "Grw Artefact", None),
             ("oi_driven_growth", "OI-driven grw", None), ("oi_level_high", "OI>25% PBT", None),
             ("oi_spike", "OI spike", None), ("oi_pct_of_pbt", "OI % of PBT", 0)]
else:
    COLS += [("coverage_3y", "Cov 3Y %", 0)]
COLS += [
    ("quality_score", "Quality raw", 1), ("growth_3y_score", "Growth 3Y raw", 1),
    ("value_score", "Value raw", 1), ("stage_3y_score", "Stage 3Y raw", 1),
    ("sector_macro_3y_score", "Sector/Macro raw", 1),
    ("revenue_growth_1y", "Rev Grw 1Y %", 1), ("roe", "ROE %", 1), ("roce", "ROCE %", 1),
    ("pe_current", "P/E", 1), ("debt_equity", "D/E", 2), ("bs_flag", "BS Gate", None),
    ("latest_qtr", "Latest Qtr", None), ("market_cap_approx", "MktCap(cr)", 0),
]

NAVY = "1F3864"; HDRC = "FFFFFF"; RED = "F4CCCC"; GREEN = "D9EAD3"; MED = "FFF2CC"; GREY = "EEEEEE"
# signal band styling: tint fill + band ink, index-aligned to the lib's DOT_COLOURS ramp
SIG_FILL = {0: ("D2EDE0", "1E9E6A"), 1: ("EAF6F0", "3E8E6E"),
            2: ("FCEBCB", "92400E"), 3: ("FBE3E0", "E0402F")}
_WORD_TO_BAND = {w: i for i, w in enumerate(F.WORDS[F.DEFAULT_WORDS])}
thin_side = Side(style="thin", color="D9D9D9")
border = Border(bottom=thin_side)

wb = Workbook(); ws = wb.active; ws.title = "All Scores (750)"
ws["A1"] = "IONIC — Nifty-750 Quant Scorecard (v8, five signals + v2 thin-history fix)"
ws["A1"].font = Font(name="Georgia", size=15, bold=True, color=NAVY)
n = len(df)
n_sell = int((df.get("recommendation_v3", pd.Series(dtype=str)) == "Sell").sum()) if has_v3 else 0
# Trim is an ELIGIBILITY, not a call: 40-50 permits a trim if the weight warrants it, and an analyst
# Sell above the bar permits one too, but neither IS a trim. The universe file has no position weights,
# so it can only flag eligibility -- the decision belongs to the book-level pass.
# fillna BEFORE astype: an empty string round-trips through CSV as NaN, and .astype(str) turns NaN into
# the four-character string "nan", which is not empty -- counting every one of the 751 names as
# trim-eligible and driving the Hold count negative.
n_trim = int((df.get("trim_eligible_v3", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
              != "").sum()) if has_v3 else 0
ws["A2"] = (f"As of {AS_OF}  |  {n} names  |  Signals = the client-deck five, quartile bands against "
            f"this universe  |  Call: Sell below 40 ({n_sell}), Hold at or above 40. {n_trim} Holds "
            f"are TRIM-ELIGIBLE (score 40-50 needs weight >2.5%, or an analyst Sell) -- eligibility, "
            f"not an instruction; the book-level pass decides  |  v3 = thin-history corrected (1y-sibling and "
            f"listing-price substitution, no withdrawals)  |  Ionic = base blend + forward adjustment "
            f"(growth leg banded on the analyst's expected EPS growth, as v1; plus the conviction leg), "
            f"capped at 5 and 95  |  Signals are trailing pillar ranks, no forward data  |  "
            f"INTERNAL RESEARCH — not investment advice.")
ws["A2"].font = Font(name="Bahnschrift", size=9, italic=True, color="666666")
HDR_ROW = 4

hdr_fill = PatternFill("solid", fgColor=NAVY)
for j, (_, label, _) in enumerate(COLS, 1):
    c = ws.cell(HDR_ROW, j, label)
    c.font = Font(name="Bahnschrift", size=9, bold=True, color=HDRC)
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, row in df.iterrows():
    r = HDR_ROW + 1 + i
    for j, (col, _, dec) in enumerate(COLS, 1):
        v = row.get(col, "")
        if pd.isna(v):
            v = ""
        elif dec is not None:
            try:
                v = round(float(v), dec)
            except (ValueError, TypeError):
                pass
        cell = ws.cell(r, j, v)
        cell.font = Font(name="Bahnschrift", size=9)
        cell.border = border
        if col.startswith("sig::"):
            band = _WORD_TO_BAND.get(v)
            cell.alignment = Alignment(horizontal="center")
            if band is not None:
                fill, ink = SIG_FILL[band]
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(name="Bahnschrift", size=9, bold=True, color=ink)
            else:
                cell.font = Font(name="Bahnschrift", size=9, color="999999")
        elif col in ("ionic_score_v3", "final_score_3y", "final_score_1y", "final_score_3y_v3",
                     "final_score_1y_v3", "recommendation_overall", "recommendation_v3"):
            cell.font = Font(name="Bahnschrift", size=9, bold=True)
    # colour the call cells by looking their column up, rather than by a hardcoded index -- the column
    # order shifts with has_v3 and a fixed index would silently paint the wrong column
    for col in ("recommendation_v3",):
        ci = next((k + 1 for k, (c0, _, _) in enumerate(COLS) if c0 == col), None)
        if ci is None:
            continue
        call = str(row.get(col, ""))
        cc = ws.cell(r, ci)
        cc.alignment = Alignment(horizontal="center")
        if call == "Sell":
            cc.fill = PatternFill("solid", fgColor=RED)
        elif call.startswith("Hold (Trim"):
            cc.fill = PatternFill("solid", fgColor=MED)
        elif call == "Hold":
            cc.fill = PatternFill("solid", fgColor=GREEN)
    if has_v3:
        hc = str(row.get("history_class", ""))
        ci = next((k + 1 for k, (c0, _, _) in enumerate(COLS) if c0 == "history_class"), None)
        if ci and hc in ("<1y", "1-2y"):
            ws.cell(r, ci).fill = PatternFill("solid", fgColor=RED if hc == "<1y" else MED)

ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(COLS))}{HDR_ROW + n}"
widths = {"Symbol": 12, "Sector": 22, "Call": 9, "Sector & Flows": 12,
          "Fwd EPS Grw % (analyst)": 10, "Imputation used": 26, "Latest Qtr": 11,
          "Listing-ret pctile": 9, "History": 8}
for j, (_, label, _) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(j)].width = widths.get(label, 9.5)
ws.row_dimensions[HDR_ROW].height = 28

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("source:", os.path.basename(src))
print("rows:", n, "cols:", len(COLS), "fwd EPS estimates joined:",
      int(df["fwd_eps_growth_pct"].notna().sum()))
if has_v3:
    # Holds and trim-eligibles OVERLAP by design: every trim-eligible name is a Hold. Reporting them
    # as three disjoint buckets was what produced "Hold -198".
    print(f"calls: Sell {n_sell} | Hold {n - n_sell}  (of which {n_trim} trim-eligible)")
print("SAVED:", OUT)
