"""
ionic_style.py - Ionic Wealth Excel house style + de-AI-ification text pass.
Implements 00_GOVERNANCE/STYLE_GUIDE.md for spreadsheet deliverables:
Georgia body / Bahnschrift headings, the firm 6-color palette, horizontal-rule
tables (no vertical rules), and detell() - the mechanical banned-tells scrub
(em/double/spaced dashes, Tier-1 words, hollow intensifiers, filler phrases).
Used by build_client_excel.py and build_analyst_excel.py.
"""
import re
from openpyxl.styles import Font, PatternFill, Border, Side

# ---- firm palette (STYLE_GUIDE.md section c) ----
INK = "1C1C1A"
NAVY = "1F3A5D"
GOLD = "B08D57"
TEAL = "2E6E62"
RUST = "A34A28"
STONE = "5F5E57"

BODY_FONT_NAME = "Georgia"
HEAD_FONT_NAME = "Bahnschrift"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name=HEAD_FONT_NAME, size=10.5, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=HEAD_FONT_NAME, size=15, bold=True, color=NAVY)
SUB_FONT = Font(name=BODY_FONT_NAME, size=9, italic=True, color=STONE)
BODY_FONT = Font(name=BODY_FONT_NAME, size=10.5, color=INK)
BODY_BOLD = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color=INK)
SMALL_FONT = Font(name=BODY_FONT_NAME, size=10, color=INK)
SMALL_BOLD = Font(name=BODY_FONT_NAME, size=10, bold=True, color=INK)

# recommendation tints (low-saturation, from the firm palette)
SELL_FILL = PatternFill(start_color="F1DFD7", end_color="F1DFD7", fill_type="solid")
SELL_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color=RUST)
TRIM_FILL = PatternFill(start_color="F3EAD9", end_color="F3EAD9", fill_type="solid")
TRIM_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color="7A6238")
HOLD_FILL = PatternFill(start_color="DFE9E6", end_color="DFE9E6", fill_type="solid")
HOLD_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color=TEAL)
BAND_FILL = PatternFill(start_color="F5F4F2", end_color="F5F4F2", fill_type="solid")  # 4-6% stone tint, >15-row tables only

# three-line table: horizontal rules only, never vertical (STYLE_GUIDE.md section d)
_HAIR = Side(style="hair", color="C9C7C3")
_RULE = Side(style="thin", color=INK)
ROW_BORDER = Border(bottom=_HAIR)
HEADER_BORDER = Border(top=_RULE, bottom=_RULE)
LAST_ROW_BORDER = Border(bottom=_RULE)


# ---- CLIENT THEME (Principal order 2026-07-18: premium wealth-platform palette) ----
# Ionic Wealth client-facing workbooks only; internal/analyst books keep the house
# palette above. Deep blue primary, semantic green/amber/red for Hold/Trim/Sell,
# grey neutrals, purple reserved as a rare analytics accent.
C_PRIMARY = "2563EB"       # deep blue
C_PRIMARY_DARK = "1D4ED8"
C_MAST = "1E3A8A"          # masthead band (darker blue for gravitas)
C_TEXT = "111827"
C_MUTED = "6B7280"
C_BORDER = "E5E7EB"
C_SURFACE = "F9FAFB"
C_CARD = "EFF6FF"          # blue-50 card tint
C_ACCENT = "6D28D9"        # purple, sparing use (analytics highlights)

C_HEADER_FILL = PatternFill(start_color=C_PRIMARY_DARK, end_color=C_PRIMARY_DARK, fill_type="solid")
C_MAST_FILL = PatternFill(start_color=C_MAST, end_color=C_MAST, fill_type="solid")
C_SURFACE_FILL = PatternFill(start_color=C_SURFACE, end_color=C_SURFACE, fill_type="solid")
C_CARD_FILL = PatternFill(start_color=C_CARD, end_color=C_CARD, fill_type="solid")
C_HEADER_FONT = Font(name=HEAD_FONT_NAME, size=10.5, bold=True, color="FFFFFF")
C_MAST_FONT = Font(name=HEAD_FONT_NAME, size=16, bold=True, color="FFFFFF")
C_MAST_SUB = Font(name=BODY_FONT_NAME, size=9, italic=True, color="DBEAFE")
C_TITLE_FONT = Font(name=HEAD_FONT_NAME, size=14, bold=True, color=C_PRIMARY_DARK)
C_KPI_NUM = Font(name=HEAD_FONT_NAME, size=19, bold=True, color=C_PRIMARY_DARK)
C_KPI_LABEL = Font(name=HEAD_FONT_NAME, size=8, bold=True, color=C_MUTED)
C_KPI_SUB = Font(name=BODY_FONT_NAME, size=8.5, italic=True, color=C_MUTED)
C_BODY = Font(name=BODY_FONT_NAME, size=10.5, color=C_TEXT)
C_BODY_BOLD = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color=C_TEXT)
C_SUB = Font(name=BODY_FONT_NAME, size=9, italic=True, color=C_MUTED)
C_SMALL = Font(name=BODY_FONT_NAME, size=10, color=C_TEXT)

# semantic recommendation colors (accessible tint/ink pairs)
C_SELL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
C_SELL_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color="B91C1C")
C_TRIM_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
C_TRIM_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color="92400E")
C_HOLD_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
C_HOLD_FONT = Font(name=BODY_FONT_NAME, size=10.5, bold=True, color="15803D")
C_BAND_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

_C_HAIR = Side(style="hair", color=C_BORDER)
_C_RULE = Side(style="thin", color="9CA3AF")
_C_CARD_EDGE = Side(style="thin", color="BFDBFE")
C_ROW_BORDER = Border(bottom=_C_HAIR)
C_HEADER_BORDER = Border(top=_C_RULE, bottom=_C_RULE)
C_LAST_ROW_BORDER = Border(bottom=_C_RULE)
C_CARD_BORDER = Border(top=_C_CARD_EDGE, bottom=_C_CARD_EDGE, left=_C_CARD_EDGE, right=_C_CARD_EDGE)


# ---- de-AI-ification text pass (STYLE_GUIDE.md sections A.1-A.5, mechanical subset) ----
_WORD_SWAPS = [
    (r"\butilize(s|d)?\b", lambda m: "use" + (m.group(1) or "").replace("s", "s").replace("d", "d")),
    (r"\bin order to\b", "to"),
    (r"\bdue to the fact that\b", "because"),
    (r"\bserves as\b", "is"),
    (r"\bboasts\b", "has"),
    (r"\bmeticulous(ly)?\b", lambda m: "careful" + ("ly" if m.group(1) else "")),
    (r"\bseamless(ly)?\b", lambda m: "smooth" + ("ly" if m.group(1) else "")),
    (r"\bdelve(s)? into\b", lambda m: "dig" + (m.group(1) or "") + " into"),
    (r"\brobust\b", "strong"),
    (r"\bcomprehensive\b", "full"),
    (r"\bholistic(ally)?\b", lambda m: "full" + ("y" if m.group(1) else "")),
    (r"\bpivotal\b", "key"),
    (r"\bunderscores\b", "shows"),
    (r"\bit is important to note that\s+", ""),
    (r"\bit'?s worth noting that\s+", ""),
    (r"\bnotably,\s+", ""),
    (r"\bmoreover,\s+", ""),
    (r"\bfurthermore,\s+", ""),
    # bare mid-sentence forms (no comma): swap, never delete (deleting breaks grammar)
    (r"\bmost notably\b", "especially"),
    (r"\bnotably\b", "especially"),
    (r"\bmoreover\b", "also"),
    (r"\bfurthermore\b", "also"),
    # hollow intensifiers (A.4): genuine/genuinely, truly — handle punctuation-adjacent too
    (r"\s*\bgenuinely\b", ""),
    (r"\s*\bgenuine\b", ""),
    (r"\s*\btruly\b", ""),
]


def detell(text):
    """Mechanical banned-tells scrub. Conservative: word-boundary swaps and dash
    normalization only - never touches numbers, tickers, or file paths."""
    if not isinstance(text, str) or not text:
        return text
    t = text
    # dashes: em dash, double hyphen, spaced single hyphen used as a dash
    t = re.sub(r"\s*—\s*", ", ", t)
    t = re.sub(r"\s+--\s+", ", ", t)
    t = re.sub(r"(?<=[a-zA-Z\)%])\s-\s(?=[a-zA-Z\(])", ", ", t)
    for pat, rep in _WORD_SWAPS:
        t = re.sub(pat, rep, t, flags=re.IGNORECASE)
    # cleanup artifacts
    t = re.sub(r"\s{2,}", " ", t)
    t = re.sub(r"\s+([,.;:])", r"\1", t)
    t = re.sub(r",\s*,", ",", t)
    t = re.sub(r"([.!?]\s+)([a-z])", lambda m: m.group(1) + m.group(2).upper(), t)  # re-capitalize after sentence break
    t = t.strip()
    if t and t[0].islower():
        t = t[0].upper() + t[1:]
    return t
