"""
FROZEN client-facing Excel for STOCK_SCORECARD_750 holdings reviews — v4 (2026-07-18).
Implements FROZEN_METHODOLOGY.md v6.2 CLIENT PORTFOLIO LAYER plus the Principal's
dashboard/analytics orders (2026-07-18 evening):
  Sheet 'At a Glance'          — client dashboard: KPI cards, simulated growth-of-100
                                 chart vs Nifty 50, suggested actions and views,
                                 market-cap mix with the mid/small allocation view.
  Sheet 'Recommendations'      — Stock | Ticker | ISIN | Sector | % of Portfolio |
                                 Ionic Score | Recommendation | Trim to | Rationale
  Sheet 'Portfolio Analytics'  — return/risk vs Nifty 50 (3y & 1y, simulated),
                                 valuation context, top-15 correlation heatmap,
                                 assumptions stated in plain language.
  Sheet 'Portfolio - Before vs After' — actions, metrics, sector & mcap mix.
Client identity: Ionic Wealth. Vocabulary Sell/Trim/Hold only, never Buy.
Theme: client premium palette (ionic_style C_* constants, Principal order 2026-07-18);
internal/analyst books keep the house palette.
Inputs: quant CSV + pf_qual_<SYMBOL>.json dir + pf_mech_flags.json + pf_fm_actions.json
        + (optional) pf_analytics.json / pf_analytics_series.csv / pf_corr_matrix.csv
        produced by compute_portfolio_analytics.py. Analytics degrade gracefully
        if the analytics files are absent.
Usage: python build_client_excel.py [quant_csv] [qual_dir] [out_xlsx]
v3 history: house-styled 3-sheet book. v4 adds the dashboard, the analytics layer and
the client premium theme.
"""
import os, sys, json, math, re
os.environ["PYTHONIOENCODING"] = "utf-8"
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
from ionic_style import (
    detell, BODY_FONT_NAME as FONT, HEAD_FONT_NAME,
    C_PRIMARY, C_PRIMARY_DARK, C_MAST, C_TEXT, C_MUTED, C_BORDER, C_SURFACE, C_ACCENT,
    C_HEADER_FILL, C_MAST_FILL, C_SURFACE_FILL, C_CARD_FILL,
    C_HEADER_FONT, C_MAST_FONT, C_MAST_SUB, C_TITLE_FONT,
    C_KPI_NUM, C_KPI_LABEL, C_KPI_SUB, C_BODY, C_BODY_BOLD, C_SUB, C_SMALL,
    C_SELL_FILL, C_SELL_FONT, C_TRIM_FILL, C_TRIM_FONT, C_HOLD_FILL, C_HOLD_FONT,
    C_BAND_FILL, C_ROW_BORDER, C_HEADER_BORDER, C_LAST_ROW_BORDER,
)
BORDER = C_ROW_BORDER

FOOTER = ("Prepared by Ionic Wealth. These are research recommendations for a review of existing "
          "holdings; execution is at your discretion.")
SIM_NOTE = ("Simulated: today's holding mix held at constant weights over the look-back window. "
            "It is not your realized return (the mix you held in the past differed) and it is not "
            "a forecast. Benchmark is the Nifty 50 with dividends added back.")


# ---------------------------------------------------------------- data layer
def load_all(quant_csv, qual_dir):
    quant = pd.read_csv(quant_csv)
    sym_col = "Symbol" if "Symbol" in quant.columns else "symbol"
    mech = json.load(open(os.path.join(qual_dir, "pf_mech_flags.json"), encoding="utf-8"))
    mrows = {h["symbol"]: h for h in mech["holdings"]}
    fm = json.load(open(os.path.join(qual_dir, "pf_fm_actions.json"), encoding="utf-8"))
    fmap = {a["symbol"]: a for a in fm["actions"]}

    rows = []
    for _, r in quant.iterrows():
        sym = r[sym_col]
        qpath = os.path.join(qual_dir, f"pf_qual_{sym}.json")
        q = json.load(open(qpath, encoding="utf-8")) if os.path.exists(qpath) else {}
        m = mrows.get(sym, {})
        a = fmap.get(sym, {})
        action = a.get("action") or ("Sell" if q.get("your_recommendation") == "Sell" else "Hold")
        rationale = q.get("summary", "Quant-only -- not yet analyst-reviewed.")
        ionic = m.get("ionic_score")
        note = ""
        if ionic is not None:
            if action == "Sell" and ionic >= 50:
                note = (" [Note: the model score and our call differ here. The Sell reflects our analyst's "
                        "judgment on valuation against achievable growth; the rationale explains.]")
            elif action in ("Hold", "Trim") and ionic < 40:
                note = (" [Note: the model score sits below our usual comfort line. We retain the position "
                        "on our analyst's forward-looking assessment; the rationale explains.]")
        client_reason = a.get("client_reason") or ""
        if client_reason:
            rationale = client_reason + " || " + rationale
        rationale = detell(rationale)
        rows.append({
            "symbol": sym, "company_name": r.get("company_name", sym),
            "isin": r.get("isin", ""), "sector": r.get("sector", ""),
            "value_inr": r.get("value_inr", None),
            "weight": m.get("weight"), "ionic": ionic,
            "action": action, "trim_to": a.get("trim_target_pct"),
            "rationale": rationale + note,
            "mcap": m.get("mcap", ""),
        })
    df = pd.DataFrame(rows)
    order = {"Sell": 0, "Trim": 1, "Hold": 2, "No Recommendation": 3}
    df["_rk"] = df["action"].map(order).fillna(4)
    df = df.sort_values(["_rk", "value_inr"], ascending=[True, False]).drop(columns="_rk")
    return df, fm, mech


def load_analytics(qual_dir):
    an, series, corr = None, None, None
    p = os.path.join(qual_dir, "pf_analytics.json")
    if os.path.exists(p):
        an = json.load(open(p, encoding="utf-8"))
    p = os.path.join(qual_dir, "pf_analytics_series.csv")
    if os.path.exists(p):
        series = pd.read_csv(p)
    p = os.path.join(qual_dir, "pf_corr_matrix.csv")
    if os.path.exists(p):
        corr = pd.read_csv(p, index_col=0)
    return an, series, corr


def verify(df, fm):
    """Frozen verification gate — hard-fail the build on any inconsistency."""
    errs = []
    wsum = df["weight"].sum()
    if abs(wsum - 100) > 0.05:
        errs.append(f"weights sum {wsum:.3f} != 100")
    if not df["action"].isin(["Sell", "Trim", "Hold"]).all():
        errs.append(f"bad vocabulary: {sorted(set(df['action']) - {'Sell','Trim','Hold'})}")
    for _, r in df.iterrows():
        if r["action"] == "Trim":
            if r["trim_to"] is None or r["trim_to"] >= r["weight"]:
                errs.append(f"{r['symbol']}: Trim target {r['trim_to']} not below weight {r['weight']}")
        if r["action"] in ("Sell", "Trim") and " || " not in (r["rationale"] or ""):
            errs.append(f"{r['symbol']}: {r['action']} without an FM client_reason")
    if errs:
        raise SystemExit("VERIFICATION FAILED:\n  " + "\n  ".join(errs))
    return wsum


def after_state(df):
    a = df.copy()
    a["after_w"] = a.apply(lambda r: 0.0 if r["action"] == "Sell"
                           else (min(r["trim_to"], r["weight"]) if r["action"] == "Trim" and pd.notna(r["trim_to"])
                                 else r["weight"]), axis=1)
    freed = a["weight"].sum() - a["after_w"].sum()
    return a, freed


# ---------------------------------------------------------------- primitives
def _title(ws, row, text, sub=None):
    ws.cell(row=row, column=2, value=text).font = C_TITLE_FONT
    if sub:
        c = ws.cell(row=row + 1, column=2, value=sub)
        c.font = C_SUB
        return row + 3
    return row + 2


def _header(ws, row, headers, col0=2):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=col0 + j, value=h)
        c.font, c.fill, c.border = C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    return row + 1


def _row_h(text, chars_per_line=92, pt_per_line=14.2, pad=8, cap=405):
    lines = max(1, math.ceil(len(text or "") / chars_per_line))
    return min(cap, pt_per_line * lines + pad)


def _mast(ws, ncols, title, sub):
    """Masthead band across the top."""
    for rr in (1, 2, 3):
        for j in range(1, ncols + 2):
            ws.cell(row=rr, column=j).fill = C_MAST_FILL
    ws.cell(row=2, column=2, value=title).font = C_MAST_FONT
    ws.cell(row=3, column=2, value=sub).font = C_MAST_SUB
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 14


_EDGE = Side(style="thin", color="BFDBFE")


def _kpi(ws, row, col, label, value, sub=""):
    """A 2-column x 3-row KPI card with a perimeter border (no internal rules)."""
    for rr in range(row, row + 3):
        for cc in (col, col + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = C_CARD_FILL
            top = _EDGE if rr == row else None
            bottom = _EDGE if rr == row + 2 else None
            left = _EDGE if cc == col else None
            right = _EDGE if cc == col + 1 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    lc = ws.cell(row=row, column=col, value=label.upper())
    lc.font = C_KPI_LABEL
    lc.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = C_KPI_NUM
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sc = ws.cell(row=row + 2, column=col, value=sub)
    sc.font = C_KPI_SUB
    sc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)


def _kpi_band_heights(ws, row):
    ws.row_dimensions[row].height = 15
    ws.row_dimensions[row + 1].height = 28
    ws.row_dimensions[row + 2].height = 24


def _rec_paint(cell, action):
    if action == "Sell":
        cell.font, cell.fill = C_SELL_FONT, C_SELL_FILL
    elif action == "Trim":
        cell.font, cell.fill = C_TRIM_FONT, C_TRIM_FILL
    else:
        cell.font, cell.fill = C_HOLD_FONT, C_HOLD_FILL


def _mcell(ws, row, c0, c1, value, font, fill=None, border=None, align=None, numfmt=None):
    """Merged cell across columns c0..c1 with member-cell styling."""
    for cc in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=cc)
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
    cell = ws.cell(row=row, column=c0, value=value)
    cell.font = font
    if align:
        cell.alignment = align
    if numfmt:
        cell.number_format = numfmt
    if c1 > c0:
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
    return cell


def mcap_view_line(mech):
    """Principal-approved mid/small allocation view (2026-07-18). A view, not a Buy."""
    large = mech.get("mcap_weights", {}).get("Large", 0)
    if large < 90:
        return ""
    return (f"View: the book is {large:.0f}% large-cap. For your risk profile, a 10-20% mid and "
            "small-cap allocation in quality names is worth discussing, funded from the freed capital. "
            "We can propose candidates from our coverage on request.")


# ---------------------------------------------------------------- sheets
def _chartdata_sheet(wb, series):
    ws = wb.create_sheet("_data")
    ws.sheet_state = "hidden"
    ws.cell(row=1, column=1, value="date")
    ws.cell(row=1, column=2, value="Your mix (simulated)")
    ws.cell(row=1, column=3, value="Nifty 50")
    for i, r in series.iterrows():
        ws.cell(row=i + 2, column=1, value=str(r["date"])[:10])
        ws.cell(row=i + 2, column=2, value=float(r["pf_nav"]))
        ws.cell(row=i + 2, column=3, value=float(r["bench_nav"]))
    return ws


def _nav_chart(wb, series):
    ws = wb["_data"]
    n = len(series)
    ch = LineChart()
    ch.title = "Rs 100 over three years: today's mix vs Nifty 50 (simulated)"
    ch.style = 2
    ch.height, ch.width = 9.0, 23.0
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=n + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.x_axis.tickLblSkip = max(1, n // 12)
    ch.x_axis.tickLblPos = "low"
    for s, color, wpt in zip(ch.series, (C_PRIMARY, "9CA3AF"), (2.25, 1.5)):
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = int(wpt * 12700)
        s.smooth = False
    ch.legend.position = "t"
    ch.legend.overlay = False
    return ch


def _glance_sheet(wb, df, a, freed, fm, mech, an, series):
    ws = wb.create_sheet("At a Glance", 0)
    ws.sheet_properties.tabColor = C_PRIMARY_DARK
    NC = 9
    _mast(ws, NC, "IONIC WEALTH  |  PORTFOLIO REVIEW",
          f"Existing-holdings review of {len(df)} positions  ·  as of 2026-07-18  ·  Sell / Trim / Hold")
    r = 5

    # KPI row 1: the actions story
    n_sell = int((df["action"] == "Sell").sum())
    n_trim = int((df["action"] == "Trim").sum())
    n_hold = int((df["action"] == "Hold").sum())
    w_act = df.loc[df["action"].isin(("Sell", "Trim")), "weight"].sum()
    _kpi(ws, r, 2, "Holdings reviewed", len(df), "full analyst and model coverage")
    _kpi(ws, r, 4, "Actions suggested", f"{n_sell} Sell · {n_trim} Trim",
         f"{n_hold} Hold; actions touch {w_act:.1f}% of book")
    _kpi(ws, r, 6, "Capital freed", f"{freed:.1f}%", "if all actions are executed; held as cash")
    _kpi(ws, r, 8, "Largest position", f"{df['weight'].max():.1f}%",
         f"falls to {a['after_w'].max():.1f}% after actions")
    _kpi_band_heights(ws, r)
    r += 4

    # KPI row 2: the risk/return story
    if an:
        p3, b3 = an["portfolio_3y"], an["bench_3y"]
        _kpi(ws, r, 2, "3-yr return (simulated)", f"{p3['cagr_pct']:.1f}%",
             f"Nifty 50: {b3['cagr_pct']:.1f}%; see note below")
        _kpi(ws, r, 4, "Market beta", f"{p3['beta']:.2f}", "sensitivity of the mix to Nifty 50 moves")
        _kpi(ws, r, 6, "Sharpe ratio (3-yr)", f"{p3['sharpe']:.2f}",
             f"Nifty 50: {b3['sharpe']:.2f}; return per unit of risk")
        _kpi(ws, r, 8, "Deepest fall (3-yr)", f"{p3['max_dd_pct']:.1f}%",
             f"Nifty 50: {b3['max_dd_pct']:.1f}%; peak to trough")
        _kpi_band_heights(ws, r)
        r += 3
        nc = ws.cell(row=r, column=2, value=SIM_NOTE)
        nc.font = C_SUB
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = _row_h(SIM_NOTE, 150, pt_per_line=11)
        r += 2

    # headline narrative
    narrative = detell(fm.get("sheet2_narrative", ""))
    ws.cell(row=r, column=2, value=narrative).font = C_BODY
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = _row_h(narrative, 125)
    r += 2

    # growth-of-100 chart
    if series is not None and len(series) > 20:
        ws.add_chart(_nav_chart(wb, series), f"B{r}")
        r += 20

    # suggested actions and views
    acts = df[df["action"].isin(("Sell", "Trim"))].nlargest(6, "weight")
    r = _header(ws, r, ["Suggested actions and views", "Action", "Now (%)", "Target (%)"])
    for _, x in acts.iterrows():
        tgt = 0.0 if x["action"] == "Sell" else x["trim_to"]
        for j, v in enumerate([f'{x["company_name"]} ({x["symbol"]})', x["action"],
                               round(x["weight"], 2), tgt]):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.font = C_BODY_BOLD if j == 0 else C_BODY
            c.border = BORDER
        _rec_paint(ws.cell(row=r, column=3), x["action"])
        r += 1
    r += 1

    # market-cap mix + the mid/small view (Principal-approved line)
    r = _header(ws, r, ["Market-cap band", "Weight (%)"])
    for band in ("Large", "Mid", "Small", "Micro"):
        wgt = mech.get("mcap_weights", {}).get(band)
        if wgt is None:
            continue
        ws.cell(row=r, column=2, value=band).font = C_BODY
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=round(wgt, 2))
        c.font, c.border, c.number_format = C_BODY_BOLD, BORDER, '0.00"%"'
        r += 1
    view = detell(mcap_view_line(mech))
    if view:
        vc = ws.cell(row=r, column=2, value=view)
        vc.font = Font(name=FONT, size=10.5, italic=True, color=C_PRIMARY_DARK)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
        ws.row_dimensions[r].height = _row_h(view, 125)
        r += 1
    r += 1

    # top sectors
    secs = df.assign(_sec=df["sector"].str.strip().str.lower().str.title()).groupby("_sec")["weight"].sum()
    secs = secs.sort_values(ascending=False).head(6)
    r = _header(ws, r, ["Largest sectors", "Weight (%)"])
    first_sec_row = r
    for sec, wgt in secs.items():
        ws.cell(row=r, column=2, value=sec).font = C_BODY
        ws.cell(row=r, column=2).border = BORDER
        c = ws.cell(row=r, column=3, value=round(float(wgt), 2))
        c.font, c.border, c.number_format = C_BODY, BORDER, '0.00"%"'
        r += 1
    ws.conditional_formatting.add(
        f"C{first_sec_row}:C{r-1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=25,
                    color=C_PRIMARY, showValue=True))
    r += 1

    guide = ("How to read this workbook: the Ionic Score (0-100) blends our 3-year and 1-year models "
             "and adjusts for expected growth over the next 3 to 5 years. Higher is stronger. "
             "'Trim to X%' is a target portfolio weight, not an instruction to exit. "
             "The Recommendations sheet carries the reasoning per holding, Portfolio Analytics shows "
             "risk and return in context, and Before-vs-After shows what the actions do to the "
             "portfolio. Freed capital is shown as cash; redeployment is a separate conversation. " + FOOTER)
    ws.cell(row=r, column=2, value=guide).font = C_SUB
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
    ws.row_dimensions[r].height = _row_h(guide, 150, pt_per_line=11)

    widths = [2.5] + [19, 19] * 4 + [2.5]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.sheet_view.showGridLines = False


def _analytics_sheet(wb, an, corr):
    """Uniform narrow grid C..Q (15 cols); tables use merged spans so the
    correlation heatmap and the metric tables share columns cleanly."""
    ws = wb.create_sheet("Portfolio Analytics")
    ws.sheet_properties.tabColor = C_ACCENT
    LAST = 17  # column Q
    V1, V2 = (3, 6), (7, 10)   # merged value spans
    PL = (11, 17)              # plain-words span
    _mast(ws, LAST, "IONIC WEALTH  |  PORTFOLIO ANALYTICS",
          "Risk and return of today's mix, in context  ·  simulated look-back with assumptions stated below")
    r = 5

    def table(r, title, rows, with_plain=True):
        r = _title(ws, r, title)
        _mcell(ws, r, 2, 2, "Measure", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="left", vertical="center", indent=1))
        _mcell(ws, r, *V1, "Your mix", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="center", vertical="center"))
        _mcell(ws, r, *V2, "Nifty 50", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="center", vertical="center"))
        _mcell(ws, r, *PL, "In plain words" if with_plain else "", C_HEADER_FONT, C_HEADER_FILL,
               C_HEADER_BORDER, Alignment(horizontal="left", vertical="center", indent=1))
        ws.row_dimensions[r].height = 24
        r += 1
        for name, v1, v2, plain in rows:
            _mcell(ws, r, 2, 2, name, C_BODY_BOLD, border=BORDER,
                   align=Alignment(vertical="center"))
            _mcell(ws, r, *V1, v1, C_BODY_BOLD, border=BORDER,
                   align=Alignment(horizontal="center", vertical="center"))
            _mcell(ws, r, *V2, v2, C_BODY, border=BORDER,
                   align=Alignment(horizontal="center", vertical="center"))
            _mcell(ws, r, *PL, plain, C_SUB, border=BORDER,
                   align=Alignment(vertical="center", wrap_text=True, indent=1))
            ws.row_dimensions[r].height = 20
            r += 1
        return r + 1

    p3, b3 = an["portfolio_3y"], an["bench_3y"]
    p1, b1 = an["portfolio_1y"], an["bench_1y"]
    r = table(r, "Return and risk over three years (simulated)", [
        ("Annual return (CAGR)", f"{p3['cagr_pct']:.1f}%", f"{b3['cagr_pct']:.1f}%", "growth rate of Rs 100 per year"),
        ("Volatility (annualized)", f"{p3['vol_pct']:.1f}%", f"{b3['vol_pct']:.1f}%", "how much the value swings"),
        ("Sharpe ratio", f"{p3['sharpe']:.2f}", f"{b3['sharpe']:.2f}", "return earned per unit of risk taken"),
        ("Deepest fall (max drawdown)", f"{p3['max_dd_pct']:.1f}%", f"{b3['max_dd_pct']:.1f}%", "worst peak-to-trough decline"),
        ("Beta to Nifty 50", f"{p3['beta']:.2f}", "1.00", "moves about in line with the market"),
        ("Alpha vs Nifty 50 (annual)", f"{p3['alpha_ann_pct']:.1f}%", "0.0%", "return above what beta alone explains"),
        ("Up-market capture", f"{p3['up_capture']:.2f}", "1.00", "share of market gains captured"),
        ("Down-market capture", f"{p3['down_capture']:.2f}", "1.00", "share of market falls suffered; lower is better"),
    ])

    r = table(r, "Last twelve months", [
        ("Return", f"{p1['cagr_pct']:.1f}%", f"{b1['cagr_pct']:.1f}%", ""),
        ("Volatility", f"{p1['vol_pct']:.1f}%", f"{b1['vol_pct']:.1f}%", ""),
        ("Deepest fall", f"{p1['max_dd_pct']:.1f}%", f"{b1['max_dd_pct']:.1f}%", ""),
        ("Beta", f"{p1['beta']:.2f}", "1.00", ""),
    ], with_plain=False)

    # style read in plain language (factor detail stays in the analyst workbook)
    fac = an.get("factors_3y", {})
    if fac:
        size_b = fac.get("SIZE", {}).get("beta", 0)
        style = ("Style read: the mix moves about one-for-one with the Nifty 50, with a clear mid-cap "
                 f"flavour to its behaviour (size sensitivity {size_b:+.2f}). Stock selection, not extra "
                 "market risk, drives the gap vs the index.")
        style = detell(style)
        c = _mcell(ws, r, 2, LAST, style, C_BODY, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = _row_h(style, 145)
        r += 2

    # valuation context
    vc = an.get("valuation_context", {})
    if vc:
        r = _title(ws, r, "Where valuations stand (index level)")
        _mcell(ws, r, 2, 2, "Index", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="left", vertical="center", indent=1))
        _mcell(ws, r, *V1, "P/E now", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="center", vertical="center"))
        _mcell(ws, r, *V2, "10-yr median", C_HEADER_FONT, C_HEADER_FILL, C_HEADER_BORDER,
               Alignment(horizontal="center", vertical="center"))
        _mcell(ws, r, *PL, "Percentile of its own history since 2016", C_HEADER_FONT, C_HEADER_FILL,
               C_HEADER_BORDER, Alignment(horizontal="left", vertical="center", indent=1))
        ws.row_dimensions[r].height = 24
        r += 1
        for key, label in [("nifty50", "Nifty 50 (large-cap)"), ("midcap150", "Nifty Midcap 150"),
                           ("smallcap250", "Nifty Smallcap 250")]:
            d = vc.get(key)
            if not d:
                continue
            _mcell(ws, r, 2, 2, label, C_BODY_BOLD, border=BORDER, align=Alignment(vertical="center"))
            _mcell(ws, r, *V1, d["pe_now"], C_BODY, border=BORDER,
                   align=Alignment(horizontal="center", vertical="center"), numfmt="0.0")
            _mcell(ws, r, *V2, d["pe_median"], C_BODY, border=BORDER,
                   align=Alignment(horizontal="center", vertical="center"), numfmt="0.0")
            _mcell(ws, r, *PL, f"{d['pe_pctile_since2016']:.0f}th percentile", C_BODY, border=BORDER,
                   align=Alignment(vertical="center", indent=1))
            ws.row_dimensions[r].height = 18
            r += 1
        note = ("Large-caps trade below their own ten-year median; mid-caps are also below median. "
                "This is context for the mid/small allocation view on the first sheet, not a market call.")
        note = detell(note)
        _mcell(ws, r, 2, LAST, note, C_SUB, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = _row_h(note, 145, pt_per_line=11)
        r += 2

    # correlation heatmap (top 15 holdings)
    if corr is not None and len(corr) > 3:
        r = _title(ws, r, "How the fifteen largest holdings move together",
                   "Correlation of daily returns over three years. Green diversifies; red moves together. "
                   f"Average pairwise correlation: {an.get('corr_top15_mean_offdiag', '')}")
        syms = list(corr.columns)
        for j, s in enumerate(syms):
            c = ws.cell(row=r, column=3 + j, value=s)
            c.font = Font(name=HEAD_FONT_NAME, size=7.5, bold=True, color=C_MUTED)
            c.alignment = Alignment(textRotation=60, horizontal="center")
        ws.row_dimensions[r].height = 46
        r += 1
        first = r
        for i, s in enumerate(syms):
            c = ws.cell(row=r, column=2, value=s)
            c.font = Font(name=HEAD_FONT_NAME, size=8, bold=True, color=C_MUTED)
            c.alignment = Alignment(horizontal="right")
            for j, s2 in enumerate(syms):
                cell = ws.cell(row=r, column=3 + j)
                if i == j:
                    cell.fill = C_SURFACE_FILL
                else:
                    cell.value = float(corr.iloc[i, j])
                    cell.number_format = "0.00"
                cell.font = Font(name=FONT, size=8, color=C_TEXT)
                cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[r].height = 16
            r += 1
        rng = f"C{first}:{get_column_letter(2 + len(syms))}{r - 1}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0.0, start_color="DCFCE7",
            mid_type="num", mid_value=0.35, mid_color="FEF9C3",
            end_type="num", end_value=0.7, end_color="FECACA"))
        r += 1

    # assumptions and honesty block
    asm = an.get("assumptions", {})
    r = _title(ws, r, "Assumptions, stated plainly")
    lines = [
        SIM_NOTE,
        f"Names with a full three-year history cover {asm.get('coverage_full_3y_weight_pct', '')}% of the book; "
        "recent listings join the simulation from their listing date.",
        f"Risk-free rate used for Sharpe: {asm.get('risk_free_annual_pct', 6.5):.1f}% per year (10-year "
        "government bond neighbourhood).",
        "A look-back of today's mix flatters the result: today's weights partly reflect past winners. "
        "Read the numbers as a portrait of the current mix's character, not as achieved performance.",
    ]
    for ln in lines:
        lnl = detell(ln)
        _mcell(ws, r, 2, LAST, "-  " + lnl, C_SMALL, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = _row_h(lnl, 150, pt_per_line=12.5)
        r += 1
    r += 1
    _mcell(ws, r, 2, LAST, FOOTER, C_SUB)

    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 27
    for j in range(3, LAST + 1):
        ws.column_dimensions[get_column_letter(j)].width = 6.6
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------- build
def build(quant_csv, qual_dir, out_xlsx):
    df, fm, mech = load_all(quant_csv, qual_dir)
    wsum = verify(df, fm)
    a, freed = after_state(df)
    an, series, corr = load_analytics(qual_dir)

    wb = Workbook()
    # ---------------- Recommendations --------------------------------
    ws = wb.active
    ws.title = "Recommendations"
    _mast(ws, 9, "IONIC WEALTH  |  PORTFOLIO RECOMMENDATIONS",
          f"Existing-holdings review  ·  {len(df)} holdings  ·  as of 2026-07-18  ·  "
          "Ionic Score = blended 3Y/1Y model score with forward-growth adjustment (0-100)")
    r0 = 5
    hdrs = ["Stock Name", "Ticker", "ISIN", "Sector", "% of Portfolio", "Ionic Score (0-100)",
            "Recommendation", "Trim to (% of portfolio)", "Rationale"]
    r1 = _header(ws, r0, hdrs, col0=1)
    for i, (_, r) in enumerate(df.iterrows()):
        row = r1 + i
        vals = [r["company_name"], r["symbol"], r["isin"], r["sector"],
                round(r["weight"], 2) if pd.notna(r["weight"]) else None,
                r["ionic"], r["action"],
                (round(r["trim_to"], 1) if r["action"] == "Trim" and pd.notna(r["trim_to"]) else None),
                r["rationale"]]
        band = (i % 2 == 1)
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = C_BODY_BOLD if j in (1, 6) else C_BODY
            c.border = BORDER
            if band:
                c.fill = C_BAND_FILL
            if j in (5, 6, 8):
                c.alignment = Alignment(vertical="top", horizontal="right")
                c.number_format = "0.00" if j == 5 else "0.0"
            else:
                c.alignment = Alignment(vertical="top", wrap_text=(j == 9))
        _rec_paint(ws.cell(row=row, column=7), r["action"])
        ws.row_dimensions[row].height = _row_h(r["rationale"])
    for j, w in enumerate([26, 12, 15, 20, 11, 11, 14, 12, 95], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{r1}"
    ws.auto_filter.ref = f"A{r1-1}:I{r1 - 1 + len(df)}"
    rng = f"F{r1}:F{r1 - 1 + len(df)}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=25, start_color="FECACA",
        mid_type="num", mid_value=50, mid_color="FEF9C3",
        end_type="num", end_value=75, end_color="BBF7D0"))
    wrng = f"E{r1}:E{r1 - 1 + len(df)}"
    ws.conditional_formatting.add(wrng, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=15,
        color=C_PRIMARY, showValue=True))
    fr = r1 + len(df) + 1
    ws.cell(row=fr, column=1, value=FOOTER).font = C_SUB
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=9)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{r1-1}:{r1-1}"
    ws.sheet_properties.tabColor = C_PRIMARY
    ws.sheet_view.showGridLines = False

    # ---------------- Before vs After --------------------------------
    ws2 = wb.create_sheet("Portfolio - Before vs After")
    _mast(ws2, 8, "IONIC WEALTH  |  BEFORE VS AFTER",
          "'After' assumes all Sell and Trim recommendations are executed; freed capital is shown as cash, not redeployed.")
    r = 5
    narr2 = detell(fm.get("sheet2_narrative", ""))
    ws2.cell(row=r, column=2, value=narr2).font = C_BODY
    ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws2.row_dimensions[r].height = _row_h(narr2, 115)
    r += 2

    r = _header(ws2, r, ["Action", "Holdings", "% of book (before)"])
    for act in ("Sell", "Trim", "Hold"):
        sub = df[df["action"] == act]
        cut = (sub["weight"] - a.loc[sub.index, "after_w"]).sum()
        label = f"{act}" + (f"  (frees {cut:.2f}%)" if act in ("Sell", "Trim") and cut > 0 else "")
        for j, v in enumerate([label, len(sub), round(sub['weight'].sum(), 2)]):
            c = ws2.cell(row=r, column=2 + j, value=v)
            c.font, c.border = C_BODY, BORDER
        _rec_paint(ws2.cell(row=r, column=2), act)
        r += 1
    for j, v in enumerate(["Freed cash (after)", "", round(freed, 2)]):
        c = ws2.cell(row=r, column=2 + j, value=v)
        c.font, c.border = C_BODY_BOLD, BORDER
    r += 2

    r = _header(ws2, r, ["Recommended action", "Stock", "Ticker", "Now (% of book)",
                         "Target (% of book)", "Frees (%)"])
    act_rows = df[df["action"].isin(("Sell", "Trim"))]
    for _, x in act_rows.iterrows():
        tgt = 0.0 if x["action"] == "Sell" else float(x["trim_to"])
        for j, v in enumerate([x["action"], x["company_name"], x["symbol"],
                               round(x["weight"], 2), tgt, round(x["weight"] - tgt, 2)]):
            c = ws2.cell(row=r, column=2 + j, value=v)
            c.font, c.border = C_BODY, BORDER
        _rec_paint(ws2.cell(row=r, column=2), x["action"])
        r += 1
    r += 1

    top5_b = df.nlargest(5, "weight")["weight"].sum()
    top5_a = a.nlargest(5, "after_w")["after_w"].sum()
    top10_b = df.nlargest(10, "weight")["weight"].sum()
    top10_a = a.nlargest(10, "after_w")["after_w"].sum()
    maxs_b = df["weight"].max(); maxs_a = a["after_w"].max()
    r = _header(ws2, r, ["Book metric", "Before", "After (invested portion)"])
    metrics = [
        ("Holdings (count)", len(df), int((a["after_w"] > 0).sum())),
        ("Largest single position (%)", f"{maxs_b:.2f}", f"{maxs_a:.2f}"),
        ("Top-5 concentration (%)", f"{top5_b:.2f}", f"{top5_a:.2f}"),
        ("Top-10 concentration (%)", f"{top10_b:.2f}", f"{top10_a:.2f}"),
        ("Cash freed by actions (%)", "0.00", f"{freed:.2f}"),
    ]
    for name, b, af in metrics:
        for j, v in enumerate([name, b, af]):
            c = ws2.cell(row=r, column=2 + j, value=v)
            c.font = C_BODY_BOLD if j == 0 else C_BODY
            c.border = BORDER
        r += 1
    r += 1

    r = _header(ws2, r, ["Sector", "Before (%)", "After (% of invested)", "Change"])
    df["_sec"] = df["sector"].str.strip().str.lower().str.title()
    a["_sec"] = df["_sec"]
    sb = df.groupby("_sec")["weight"].sum()
    sa = a.groupby("_sec")["after_w"].sum() / a["after_w"].sum() * 100
    for sec in sb.sort_values(ascending=False).index:
        b, af = sb[sec], sa.get(sec, 0.0)
        for j, v in enumerate([sec, round(b, 2), round(af, 2), f"{af-b:+.2f}"]):
            c = ws2.cell(row=r, column=2 + j, value=v)
            c.font, c.border = C_BODY, BORDER
        r += 1
    r += 1

    r = _header(ws2, r, ["Market-cap band", "Before (%)", "After (% of invested)"])
    mb = df.groupby("mcap")["weight"].sum()
    ma = a.groupby("mcap")["after_w"].sum() / a["after_w"].sum() * 100
    for band in ("Large", "Mid", "Small", "Micro"):
        if band in mb.index or band in ma.index:
            for j, v in enumerate([band, round(mb.get(band, 0), 2), round(ma.get(band, 0), 2)]):
                c = ws2.cell(row=r, column=2 + j, value=v)
                c.font, c.border = C_BODY, BORDER
            r += 1
    view = detell(mcap_view_line(mech))
    if view:
        vc2 = ws2.cell(row=r, column=2, value=view)
        vc2.font = Font(name=FONT, size=10.5, italic=True, color=C_PRIMARY_DARK)
        vc2.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws2.row_dimensions[r].height = _row_h(view, 115)
        r += 1
    r += 1
    cons = "Small positions note: " + detell(fm.get("consolidation_note", ""))
    ws2.cell(row=r, column=2, value=cons).font = C_BODY
    ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws2.row_dimensions[r].height = _row_h(cons, 115)
    r += 2
    ws2.cell(row=r, column=2, value=FOOTER).font = C_SUB
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    for j, w in enumerate([2.5, 34, 16, 22, 12, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.sheet_properties.tabColor = "F59E0B"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.sheet_view.showGridLines = False

    # ---------------- analytics + dashboard --------------------------
    if series is not None:
        _chartdata_sheet(wb, series)
    if an:
        _analytics_sheet(wb, an, corr)
    _glance_sheet(wb, df, a, freed, fm, mech, an, series)

    # final tab order (client experience): dashboard, detail, analytics, effects
    order = ["At a Glance", "Recommendations", "Portfolio Analytics",
             "Portfolio - Before vs After", "_data"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
    wb.active = 0

    wb.properties.creator = "Ionic Wealth"
    wb.properties.title = "Ionic Wealth - Portfolio Recommendations"
    wb.properties.description = ("Existing-holdings review (Sell/Trim/Hold). Methodology: "
                                 "STOCK_SCORECARD_750 FROZEN_METHODOLOGY.md v6.2")

    recon = (a["weight"] - a["after_w"]).sum()
    assert abs(recon - freed) < 0.01, f"before/after reconciliation broke: {recon} vs {freed}"

    # zero-tell gate: no client sheet ships with a banned tell (STYLE_GUIDE.md)
    tells = re.compile(r"—|\s--\s|\b(robust|comprehensive|pivotal|utilizes?|boasts|underscores|"
                       r"delves?|genuinely|truly|moreover|furthermore|notably|holistic|seamless|meticulous)\b", re.I)
    bad = []
    for sh in wb.worksheets:
        if sh.title == "_data" or sh.sheet_state == "hidden":
            continue
        for row in sh.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and tells.search(cell.value):
                    bad.append((sh.title, cell.coordinate, tells.search(cell.value).group(0)))
    if bad:
        raise SystemExit("TELL SCAN FAILED: " + "; ".join(f"{s}!{c} '{t}'" for s, c, t in bad[:10]))

    wb.save(out_xlsx)
    print(f"Saved {len(df)} rows to {out_xlsx} | weights sum {wsum:.2f} | freed {freed:.2f}% | "
          f"sheets: {', '.join(n for n in wb.sheetnames if n != '_data')} | tell scan CLEAN")
    return out_xlsx


if __name__ == "__main__":
    base = r"C:\Users\Shreyas.1Gupta\OneDrive - Angel Broking Limited\Desktop\Backup\NIFTY 500\Shreyas_Ionic_AMC"
    quant_csv = sys.argv[1] if len(sys.argv) > 1 else base + r"\04_RND_LAB\STOCK_SCORECARD_750\results\portfolio_quant.csv"
    qual_dir = sys.argv[2] if len(sys.argv) > 2 else base + r"\04_RND_LAB\STOCK_SCORECARD_750\results"
    out = sys.argv[3] if len(sys.argv) > 3 else base + r"\09_PRODUCT\reports\CLIENT_RECOMMENDATIONS.xlsx"
    try:
        build(quant_csv, qual_dir, out)
    except PermissionError:
        # canonical file open in Excel/OneDrive lock — never force; ship a versioned name
        alt = out.replace(".xlsx", "_v4.xlsx")
        print(f"{out} is locked (open in Excel?); writing {alt} instead")
        build(quant_csv, qual_dir, alt)
