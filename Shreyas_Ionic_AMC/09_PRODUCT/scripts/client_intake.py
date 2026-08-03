# -*- coding: utf-8 -*-
"""client_intake.py — NDPMS deck intake layer (Principal ruling 2026-07-26: primary
intake = NSDL CAS; exceptions handled by scraping/typing from whatever the client gives).

Flow
  1. Holdings arrive as a CSV/XLSX extracted from the CAS (or typed by the RM for
     exception cases). Expected columns (case-insensitive, extras ignored):
        type      : EQ | MF
        name      : security / scheme name as printed
        isin      : optional but strongly preferred (exact matching)
        units     : quantity / units held
        value_inr : current value in INR (if absent, price*units must be derivable)
     A raw NSDL CAS PDF parser slots in here once a sample statement is provided —
     the PDF's holdings tables map to exactly this frame.
  2. Equity rows are matched to the scored 750 universe (portfolio_quant.csv) by
     ISIN first, then normalized-name prefix; MF rows go to fund_ctx_adapter for
     QFRA-wired calls. Anything unmatched lands in exceptions.csv for the RM —
     nothing silently dropped, nothing fabricated (D-035).
  3. The client profile (goals, holding ages/costs, family, meeting history) is a
     JSON file the RM maintains; see PROFILE_TEMPLATE below. intake emits a single
     client_ctx.json the deck's data layer consumes.

Usage
  python client_intake.py --holdings <file.csv|xlsx> --profile <profile.json>
                          --out <client_dir>
"""
import os
import re
import sys
import csv
import json
import argparse

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.abspath(os.path.join(HERE, "..", "..", "04_RND_LAB",
                                       "STOCK_SCORECARD_750", "results"))
sys.path.insert(0, os.path.abspath(os.path.join(HERE, "..", "pr_template")))
from lib.mf_mapping import validate_holdings, resolve_scheme_rename  # noqa: E402

# the four personalization blocks every real deck carries (Principal 2026-07-26)
PROFILE_TEMPLATE = {
    "client": {"name": "", "code": "", "account_type": "NDPMS (Non-Discretionary)",
               "profile": "Moderate", "horizon": "", "construction": "Core-satellite"},
    "goals": [
        {"name": "e.g. Education 2031", "target_inr": 0, "by_year": 0, "note": ""},
    ],
    "holdings_meta": {
        # per symbol/scheme: acquisition info the CAS does not carry
        # "RELIANCE": {"holding_years": 6.2, "cost_inr": 0},
    },
    "family": {"members": [], "structure_note": ""},
    "meeting_history": [
        # newest first; see modules/since_last_review.py for the rendered shape
        # {"date": "2026-04-28", "summary": "...", "actions": [{"action": "...",
        #   "owner": "Client", "status": "Done"}]}
    ],
}


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _load_holdings(path):
    rows = []
    if path.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        head = None
        for r in ws.iter_rows(values_only=True):
            if head is None:
                head = [_norm(c) for c in r]
                continue
            rows.append({h: v for h, v in zip(head, r)})
    else:
        with open(path, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                rows.append({_norm(k): v for k, v in r.items()})
    return rows


def _load_universe():
    """symbol -> row of the scored 750; plus isin/name lookup maps."""
    by_symbol, by_isin, by_name = {}, {}, {}
    with open(os.path.join(RESULTS, "portfolio_quant.csv"), encoding="utf-8") as f:
        for r in csv.DictReader(f):
            by_symbol[r["symbol"]] = r
            if r.get("isin"):
                by_isin[r["isin"].strip().upper()] = r
            by_name[_norm(r.get("Company Name") or r["symbol"])] = r
    return by_symbol, by_isin, by_name


def _match_equity(row, by_isin, by_name):
    isin = str(row.get("isin") or "").strip().upper()
    if isin and isin in by_isin:
        return by_isin[isin], "isin"
    key = _norm(row.get("name"))
    if key in by_name:
        return by_name[key], "name-exact"
    # longest-prefix fallback (>=10 normalized chars)
    best, best_n = None, 0
    for k, v in by_name.items():
        n = 0
        while n < min(len(key), len(k)) and key[n] == k[n]:
            n += 1
        if n > best_n:
            best, best_n = v, n
    return (best, "name-prefix") if best_n >= 10 else (None, "unmatched")


def intake(holdings_path, profile_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    raw = _load_holdings(holdings_path)
    profile = json.load(open(profile_path, encoding="utf-8")) if os.path.exists(profile_path) else dict(PROFILE_TEMPLATE)
    _, by_isin, by_name = _load_universe()

    # row-bleed / CAS-corruption check (2026-08-02, Talaulikar build: a stock's name
    # field had an entirely different holding's row content silently concatenated
    # onto it during extraction). Flag, never silently trust or drop.
    row_warnings = validate_holdings(raw)
    if row_warnings:
        print(f"intake: {len(row_warnings)} row(s) flagged by data-quality check (see "
              f"row_warnings.json in {out_dir}) -- review before trusting these rows")
        json.dump({str(i): w for i, w in row_warnings.items()},
                  open(os.path.join(out_dir, "row_warnings.json"), "w", encoding="utf-8"),
                  indent=2, ensure_ascii=False)

    equity, mf, exceptions = [], [], []
    for row in raw:
        typ = str(row.get("type") or "").strip().upper()
        val = row.get("valueinr") or row.get("value") or 0
        try:
            val = float(val)
        except Exception:
            exceptions.append({**row, "reason": "value not numeric"})
            continue
        if typ == "MF":
            # resolve known AMFI scheme renames so this and future clients' holdings
            # match our fund-quality frameworks on the first pass (see lib/mf_mapping.py)
            resolved_name = resolve_scheme_rename(row.get("name") or "")
            mf.append({"name": resolved_name, "raw_name": row.get("name"),
                       "value_inr": val, "units": row.get("units")})
        else:
            hit, how = _match_equity(row, by_isin, by_name)
            if hit is None:
                exceptions.append({**row, "reason": "no match in scored 750 universe"})
                continue
            meta = (profile.get("holdings_meta") or {}).get(hit["symbol"], {})
            equity.append({"symbol": hit["symbol"], "name": row.get("name"),
                           "value_inr": val, "match": how,
                           "holding_years": meta.get("holding_years"),
                           "cost_inr": meta.get("cost_inr")})

    total = sum(e["value_inr"] for e in equity) + sum(f["value_inr"] for f in mf)
    for e in equity:
        e["weight_pct"] = round(100.0 * e["value_inr"] / total, 2) if total else 0
    for f in mf:
        f["weight_pct"] = round(100.0 * f["value_inr"] / total, 2) if total else 0

    # is_demo=False explicit here (2026-07-28) -- every module in the deck engine defaults
    # ctx.get("is_demo", False) now (was accidentally True firm-wide, which would silently
    # print "illustrative synthetic" disclaimers on a real client's deck if this key were ever
    # missing). Anything going through the real intake pipeline is never a demo -- make that a
    # fact stamped at the single point of truth, not an assumption left to each new client's
    # data/<client>.py author to remember.
    ctx = {"profile": profile, "equity": equity, "mf": mf, "is_demo": False,
           "totals": {"grand_inr": total, "n_equity": len(equity), "n_mf": len(mf),
                      "n_exceptions": len(exceptions)}}
    out_ctx = os.path.join(out_dir, "client_ctx.json")
    json.dump(ctx, open(out_ctx, "w", encoding="utf-8"), indent=2, ensure_ascii=False)

    out_exc = os.path.join(out_dir, "exceptions.csv")
    if exceptions:
        keys = sorted({k for e in exceptions for k in e})
        with open(out_exc, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(exceptions)
    print(f"intake: {len(equity)} equity matched, {len(mf)} MF rows, "
          f"{len(exceptions)} exceptions -> {out_ctx}")
    if exceptions:
        print(f"RM ACTION NEEDED: review {out_exc} (unmatched/invalid rows are never "
              f"silently dropped)")
    return ctx


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--holdings", required=True)
    ap.add_argument("--profile", default="")
    ap.add_argument("--out", default=".")
    ap.add_argument("--emit-template", action="store_true",
                    help="write profile_template.json beside --out and exit")
    a = ap.parse_args()
    if a.emit_template:
        p = os.path.join(a.out, "profile_template.json")
        json.dump(PROFILE_TEMPLATE, open(p, "w", encoding="utf-8"), indent=2)
        print("template ->", p)
    else:
        intake(a.holdings, a.profile, a.out)
